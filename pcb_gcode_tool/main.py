from __future__ import annotations

import argparse
import base64
import json
import math
import os
import re
import select
import socket
import subprocess
import sys
import tempfile
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from PyQt6.QtCore import QPointF, QRectF, Qt, QTimer
from PyQt6.QtGui import (QAction, QColor, QFont, QPainter, QPainterPath, QPen,
                         QPolygonF, QTextCursor)
from PyQt6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDialog, QDoubleSpinBox, QFileDialog, QFormLayout,
    QFrame, QGridLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit,
    QMainWindow, QMessageBox, QPlainTextEdit, QPushButton, QScrollArea,
    QSizePolicy, QSplitter, QStackedWidget, QStatusBar, QStyle, QTableWidget,
    QTableWidgetItem, QToolBar, QVBoxLayout, QWidget,
)

try:
    from gerber_parser import GerberPad, parse_gerber_file
except ImportError:
    from .gerber_parser import GerberPad, parse_gerber_file


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = ROOT.parent / "PickAndPlace_PCB_PCB_1048_looper_2_2025_10_07_2_2026_08_27.xlsx"
DEVICE_CONFIG_PATH = ROOT / "device_connection.json"
DISCOVERY_PORT = 8267
BRIDGE_PORT = 8266
UPLOAD_CHUNK_SIZE = 48
DEFAULT_DEVICE_IP = "192.168.4.1"  # AP 模式固定 IP
DEFAULT_DEVICE_NAME = "BanPCBTool"  # 模块热点名称前缀
DEFAULT_AP_PASSWORD = "12345678"    # Ban-IOT 协议统一 AP 密码


def _pump_sleep(seconds: float) -> None:
    """阻塞等待期间保持 GUI 响应：边睡边处理 Qt 事件。

    用在扫描 WiFi、切换热点轮询等较长阻塞等待中，避免界面卡死、
    日志不刷新。非 Qt 环境下等价于普通 sleep。
    """
    deadline = time.monotonic() + seconds
    while time.monotonic() < deadline:
        try:
            QApplication.processEvents()
        except RuntimeError:
            pass
        time.sleep(0.05)


def scan_local_wifi() -> list[str]:
    """使用本机网卡扫描周围 WiFi。

    优先调用 Windows WLAN API（wlanapi.dll）并主动触发一次扫描，
    可以拿到附近全部可见 WiFi；netsh 在无线网卡已连接时通常
    只列出当前连接的网络，仅作为回退。
    """
    if sys.platform != "win32":
        return []
    # WLAN API 与 netsh 结果取并集：有的网卡驱动 netsh 只返回
    # 当前连接网络，而 WLAN API 能列出全部；反之亦然，两者互补。
    names: list[str] = []
    for fn in (_scan_wlanapi, _scan_netsh):
        try:
            for name in fn():
                if name and name not in names:
                    names.append(name)
        except Exception:
            continue
    return names


def _scan_netsh() -> list[str]:
    try:
        output = subprocess.check_output(
            ["netsh", "wlan", "show", "networks"],
            timeout=15.0,
            encoding="utf-8",
            errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except (OSError, subprocess.SubprocessError):
        return []
    ssids: list[str] = []
    for line in output.splitlines():
        stripped = line.strip()
        if stripped.lower().startswith("ssid") and ":" in stripped:
            name = stripped.split(":", 1)[1].strip()
            if name and name not in ssids:
                ssids.append(name)
    return ssids


def _scan_wlanapi() -> list[str]:
    """通过 wlanapi.dll 触发无线网卡扫描并读取完整可用网络列表。"""
    import ctypes
    from ctypes import wintypes

    class WLAN_INTERFACE_INFO(ctypes.Structure):
        _fields_ = [
            ("InterfaceGuid", ctypes.c_ubyte * 16),
            ("strInterfaceDescription", wintypes.WCHAR * 256),
            ("isState", wintypes.DWORD),
        ]

    class WLAN_INTERFACE_INFO_LIST(ctypes.Structure):
        _fields_ = [
            ("dwNumberOfItems", wintypes.DWORD),
            ("dwIndex", wintypes.DWORD),
            ("InterfaceInfo", WLAN_INTERFACE_INFO * 1),
        ]

    class WLAN_AVAILABLE_NETWORK(ctypes.Structure):
        _fields_ = [
            ("ProfileName", wintypes.WCHAR * 256),
            ("SSIDLength", wintypes.DWORD),
            ("SSID", ctypes.c_ubyte * 32),
            ("BssType", wintypes.DWORD),
            ("AuthAlgorithm", wintypes.DWORD),
            ("CipherAlgorithm", wintypes.DWORD),
            ("Flags", wintypes.DWORD),
            ("SignalQuality", wintypes.DWORD),
            ("NumSecurityAlgorithms", wintypes.DWORD),
            ("SecurityAlgorithms", wintypes.DWORD * 1),
            ("PhyTypeNumber", wintypes.DWORD),
            ("PhyTypes", wintypes.DWORD * 1),
        ]

    class WLAN_AVAILABLE_NETWORK_LIST(ctypes.Structure):
        _fields_ = [
            ("dwNumberOfItems", wintypes.DWORD),
            ("dwIndex", wintypes.DWORD),
            ("Network", WLAN_AVAILABLE_NETWORK * 1),
        ]

    try:
        wlanapi = ctypes.WinDLL("wlanapi", use_last_error=True)
    except OSError:
        return []

    wlanapi.WlanOpenHandle.argtypes = [
        wintypes.DWORD, ctypes.c_void_p,
        ctypes.POINTER(wintypes.DWORD), ctypes.POINTER(wintypes.HANDLE)]
    wlanapi.WlanOpenHandle.restype = wintypes.DWORD
    wlanapi.WlanEnumInterfaces.argtypes = [
        wintypes.HANDLE, ctypes.c_void_p,
        ctypes.POINTER(ctypes.POINTER(WLAN_INTERFACE_INFO_LIST))]
    wlanapi.WlanEnumInterfaces.restype = wintypes.DWORD
    wlanapi.WlanScan.argtypes = [wintypes.HANDLE, ctypes.c_void_p] + [ctypes.c_void_p] * 3
    wlanapi.WlanScan.restype = wintypes.DWORD
    wlanapi.WlanGetAvailableNetworkList.argtypes = [
        wintypes.HANDLE, ctypes.c_void_p, wintypes.DWORD, ctypes.c_void_p,
        ctypes.POINTER(ctypes.POINTER(WLAN_AVAILABLE_NETWORK_LIST))]
    wlanapi.WlanGetAvailableNetworkList.restype = wintypes.DWORD
    wlanapi.WlanFreeMemory.argtypes = [ctypes.c_void_p]
    wlanapi.WlanCloseHandle.argtypes = [wintypes.HANDLE, ctypes.c_void_p]
    wlanapi.WlanCloseHandle.restype = wintypes.DWORD

    handle = wintypes.HANDLE()
    negotiated = wintypes.DWORD()
    if wlanapi.WlanOpenHandle(2, None, ctypes.byref(negotiated),
                              ctypes.byref(handle)) != 0:
        return []
    try:
        p_interfaces = ctypes.POINTER(WLAN_INTERFACE_INFO_LIST)()
        if wlanapi.WlanEnumInterfaces(handle, None, ctypes.byref(p_interfaces)) != 0:
            return []
        interfaces = p_interfaces.contents
        if interfaces.dwNumberOfItems == 0:
            return []
        guid = interfaces.InterfaceInfo[0].InterfaceGuid
        guid_ptr = ctypes.cast(ctypes.byref(guid), ctypes.c_void_p)
        # 主动触发一次扫描，等待几秒让网卡收集周围 WiFi
        if wlanapi.WlanScan(handle, guid_ptr, None, None, None) != 0:
            return []
        _pump_sleep(3.0)
        p_list = ctypes.POINTER(WLAN_AVAILABLE_NETWORK_LIST)()
        if wlanapi.WlanGetAvailableNetworkList(
                handle, guid_ptr, 0, None, ctypes.byref(p_list)) != 0:
            return []
        networks = p_list.contents
        networks_arr = ctypes.cast(
            ctypes.addressof(networks.Network),
            ctypes.POINTER(WLAN_AVAILABLE_NETWORK))
        ssids: list[str] = []
        for i in range(networks.dwNumberOfItems):
            network = networks_arr[i]
            length = min(network.SSIDLength, 32)
            raw = bytes(network.SSID[:length])
            try:
                name = raw.decode("utf-8", "ignore")
            except Exception:
                continue
            name = name.strip()
            if name and name.isprintable() and name not in ssids:
                ssids.append(name)
        return ssids
    finally:
        wlanapi.WlanCloseHandle(handle, None)


def current_wifi_ssid() -> str:
    """返回电脑当前连接的 WiFi 名称；未连接无线网卡时返回空串。"""
    if sys.platform != "win32":
        return ""
    try:
        output = subprocess.check_output(
            ["netsh", "wlan", "show", "interfaces"],
            timeout=10.0, encoding="utf-8", errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
    except (OSError, subprocess.SubprocessError):
        return ""
    for line in output.splitlines():
        stripped = line.strip()
        if stripped.lower().startswith("ssid") and ":" in stripped:
            name = stripped.split(":", 1)[1].strip()
            if name:
                return name
    return ""


def _netsh(args: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        ["netsh", "wlan", *args],
        capture_output=True, text=True, errors="replace", timeout=15.0,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))


def get_local_broadcast_addresses() -> list[str]:
    """Return all directed broadcast addresses the current host is on
    (e.g. 192.168.1.255 for each NIC) + 255.255.255.255 + 192.168.4.255 (module AP).
    Windows routing often silently drops 255.255.255.255 on multi-NIC PCs,
    so the directed subnet broadcasts dramatically improve discovery hit rate.
    """
    result: set[str] = {"255.255.255.255", "192.168.4.255"}
    try:
        import netifaces  # type: ignore
        for iface in netifaces.interfaces():
            for entry in netifaces.ifaddresses(iface).get(netifaces.AF_INET, []):
                ip = entry.get("addr", "")
                mask = entry.get("netmask", "")
                if not ip or not mask or ip.startswith("127."):
                    continue
                try:
                    ip_bytes = socket.inet_aton(ip)
                    mask_bytes = socket.inet_aton(mask)
                    bcast_bytes = bytes(ib | (~mb & 0xFF) for ib, mb in zip(ip_bytes, mask_bytes, strict=False))
                    result.add(socket.inet_ntoa(bcast_bytes))
                except OSError:
                    pass
        return list(result)
    except Exception:
        try:
            _, _, ips = socket.gethostbyname_ex(socket.gethostname())
            for ip in ips:
                if ip.startswith("127."):
                    continue
                parts = ip.split(".")
                if len(parts) == 4:
                    result.add(f"{parts[0]}.{parts[1]}.{parts[2]}.255")
        except (socket.herror, OSError):
            pass
        return list(result)


def get_local_lan_prefixes() -> list[str]:
    """Return list of "x.y.z" 24-bit prefixes this PC lives on (sorted by priority).
    优先选 192.168.x.x → 10.x.x.x → 172.16-31.x.x；忽略 127/169.254/192.168.4(模块AP)。
    每个前缀对应的本机完整 IP 用空格另存。返回 [("192.168.220", "192.168.220.15"), ...]。
    """
    nics: list[tuple[str, str]] = []
    try:
        import netifaces  # type: ignore
        for iface in netifaces.interfaces():
            for entry in netifaces.ifaddresses(iface).get(netifaces.AF_INET, []):
                ip = entry.get("addr", "")
                if not ip or ip.startswith("127.") or ip.startswith("169.254."):
                    continue
                parts = ip.split(".")
                if len(parts) != 4:
                    continue
                # 跳过模块 AP 默认网段（192.168.4.x）—— 除非它是唯一网卡
                if parts[0] == "192" and parts[1] == "168" and parts[2] == "4":
                    continue
                prefix = f"{parts[0]}.{parts[1]}.{parts[2]}"
                nics.append((prefix, ip))
        if nics:
            # 按常见家庭路由网段优先：192.168.x > 172.16-31 > 10.x
            def _rank(item: tuple[str, str]) -> int:
                p = item[0]
                if p.startswith("192.168."):
                    return 0
                if p.startswith("10."):
                    return 2
                a, b, _ = p.split(".", 2)
                if a == "172" and 16 <= int(b) <= 31:
                    return 1
                return 3
            nics.sort(key=_rank)
            return nics
    except Exception:
        pass
    # fallback without netifaces
    try:
        _, _, ips = socket.gethostbyname_ex(socket.gethostname())
        for ip in ips:
            if not ip or ip.startswith("127.") or ip.startswith("169.254.") or ip.startswith("192.168.4."):
                continue
            parts = ip.split(".")
            if len(parts) == 4:
                nics.append((f"{parts[0]}.{parts[1]}.{parts[2]}", ip))
    except (socket.herror, OSError):
        pass
    return nics


def is_ip_busy(ip: str, tcp_ports: tuple[int, ...] = (80, 8266), tcp_timeout: float = 0.25) -> bool:
    """快速判断一个 IP 是否被局域网内其他主机占用。
    先试 TCP 80(HTTP)/8266(桥接) 200ms 快速 SYN；都不通再试 ICMP ping（subprocess ping，800ms 超时）。
    任何一个通就算 IP 已占用。
    """
    for port in tcp_ports:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(tcp_timeout)
        try:
            s.connect((ip, port))
            return True
        except (socket.timeout, TimeoutError, OSError):
            pass
        finally:
            try:
                s.close()
            except OSError:
                pass
    # ICMP 兜底：Windows ping -n 1 -w 600
    try:
        if sys.platform.startswith("win"):
            proc = subprocess.run(
                ["ping.exe", "-n", "1", "-w", "600", "-l", "1", ip],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                timeout=3, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
            return proc.returncode == 0
        proc = subprocess.run(
            ["ping", "-c", "1", "-W", "1", ip],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=3)
        return proc.returncode == 0
    except (OSError, subprocess.SubprocessError):
        return False


def suggest_static_ip(fallback_prefix: str = "192.168.1",
                      avoid_ips: list[str] | None = None,
                      check_conflict: bool = True,
                      max_tries: int = 12) -> str:
    """推荐一个无人使用的静态 IP（对齐用户需求：192.168.xx.xx，后两段随机/不冲突）。
    - 自动获取 PC 所在的家庭网段前缀 x.y.z；拿不到才用 fallback_prefix（默认 192.168.1）
    - 主机号在 100~240 之间随机挑候选，避开 avoid_ips（saved_device.ip 等）
    - check_conflict=True 时对每个候选做 TCP 80/8266 + ICMP 冲突检测
    - 12 轮都冲突就返回最后一个候选（让用户自己改），不会抛异常
    """
    prefixes = get_local_lan_prefixes()
    if prefixes:
        chosen_prefix, _my_ip = prefixes[0]
    else:
        chosen_prefix = fallback_prefix

    avoid = set(avoid_ips or [])
    # 把本机 IP 也加到避免列表
    for _, myip in prefixes:
        avoid.add(myip)
    # 典型网关 x.y.z.1 / x.y.z.254 不选
    avoid.add(f"{chosen_prefix}.1")
    avoid.add(f"{chosen_prefix}.254")
    avoid.add(f"{chosen_prefix}.255")

    import random
    rng = random.Random()
    candidate = ""
    for _ in range(max_tries):
        host = rng.randint(100, 240)
        ip = f"{chosen_prefix}.{host}"
        if ip in avoid:
            continue
        candidate = ip
        if not check_conflict:
            return ip
        if not is_ip_busy(ip):
            return ip
    return candidate or f"{chosen_prefix}.200"


def build_wlan_profile(ssid: str, password: str | None = None) -> str:
    """生成 WLAN profile XML。
    - 有 password → WPA2-PSK/AES（模块 AP 模式，对齐 Ban-IOT）
    - 无 password → 开放热点（向后兼容）
    """
    esc_s = ssid.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    if password:
        esc_p = password.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return (
            '<?xml version="1.0"?>\n'
            '<WLANProfile xmlns="http://www.microsoft.com/networking/WLAN/profile/v1">\n'
            f"    <name>{esc_s}</name>\n"
            "    <SSIDConfig>\n"
            "        <SSID>\n"
            f"            <name>{esc_s}</name>\n"
            "        </SSID>\n"
            "    </SSIDConfig>\n"
            "    <connectionType>ESS</connectionType>\n"
            "    <connectionMode>auto</connectionMode>\n"
            "    <MSM>\n"
            "        <security>\n"
            "            <authEncryption>\n"
            "                <authentication>WPA2PSK</authentication>\n"
            "                <encryption>AES</encryption>\n"
            "                <useOneX>false</useOneX>\n"
            "            </authEncryption>\n"
            "            <sharedKey>\n"
            "                <keyType>passPhrase</keyType>\n"
            "                <protected>false</protected>\n"
            f"                <keyMaterial>{esc_p}</keyMaterial>\n"
            "            </sharedKey>\n"
            "        </security>\n"
            "    </MSM>\n"
            "</WLANProfile>"
        )
    return (
        '<?xml version="1.0"?>\n'
        '<WLANProfile xmlns="http://www.microsoft.com/networking/WLAN/profile/v1">\n'
        f"    <name>{esc_s}</name>\n"
        "    <SSIDConfig>\n"
        "        <SSID>\n"
        f"            <name>{esc_s}</name>\n"
        "        </SSID>\n"
        "    </SSIDConfig>\n"
        "    <connectionType>ESS</connectionType>\n"
        "    <connectionMode>auto</connectionMode>\n"
        "    <MSM>\n"
        "        <security>\n"
        "            <authEncryption>\n"
        "                <authentication>open</authentication>\n"
        "                <encryption>none</encryption>\n"
        "                <useOneX>false</useOneX>\n"
        "            </authEncryption>\n"
        "        </security>\n"
        "    </MSM>\n"
        "</WLANProfile>"
    )


def build_open_profile(ssid: str) -> str:
    """兼容旧代码：生成开放热点（无密码）的 WLAN profile XML。"""
    return build_wlan_profile(ssid, None)


def connect_to_hotspot(ssid: str, password: str | None = None,
                       timeout: float = 15.0) -> bool:
    """把电脑 Wi-Fi 自动切换到指定热点；连上后返回 True。
    - 有 password → WPA2-PSK（新固件 AP 模式）
    - 无 password → 开放热点（向后兼容）
    """
    if sys.platform != "win32":
        return False
    path = None
    try:
        with tempfile.NamedTemporaryFile(
                "w", suffix=".xml", delete=False, encoding="utf-8") as f:
            f.write(build_wlan_profile(ssid, password))
            path = f.name
        # 覆盖 profile 再 connect；重复 add 一般只是 warning，可忽略
        _netsh(["add", "profile", f"filename={path}", "user=current"])
        _netsh(["connect", f"name={ssid}"])
    except (OSError, subprocess.SubprocessError):
        return False
    finally:
        if path:
            try:
                os.unlink(path)
            except OSError:
                pass
    deadline = time.time() + timeout
    while time.time() < deadline:
        _pump_sleep(1.0)
        if current_wifi_ssid() == ssid:
            return True
    return False


def reconnect_wifi(ssid: str) -> None:
    """用系统已保存的 profile 回连指定 WiFi（不新建 profile，避免覆盖密码）。"""
    if sys.platform != "win32":
        return
    try:
        _netsh(["connect", f"name={ssid}"])
    except (OSError, subprocess.SubprocessError):
        pass


@dataclass
class WirelessDevice:
    name: str
    ip: str
    bridge_port: int = BRIDGE_PORT
    http_port: int = 80
    wifi: str = ""
    ap: bool = False


@dataclass
class SheetData:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass
class PointData:
    row_id: int
    raw: dict[str, Any]
    x: float
    y: float


@dataclass
class DispenseSegment:
    pad_id: int
    label: str
    start: tuple[float, float]
    end: tuple[float, float]

    @property
    def length(self) -> float:
        return math.hypot(self.end[0] - self.start[0], self.end[1] - self.start[1])


def load_workbook(path: Path) -> list[SheetData]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets: list[SheetData] = []
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                sheets.append(SheetData(sheet.title, [], []))
                continue
            headers = [str(value).strip() if value is not None else f"Column {index + 1}"
                       for index, value in enumerate(first)]
            rows: list[dict[str, Any]] = []
            for source_row in iterator:
                if not any(value is not None and str(value).strip() for value in source_row):
                    continue
                rows.append({headers[index]: value for index, value in enumerate(source_row)
                             if index < len(headers) and value is not None})
            sheets.append(SheetData(sheet.title, headers, rows))
    finally:
        workbook.close()
    if not sheets:
        raise ValueError("工作簿中没有工作表")
    return sheets


def parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[+-]?(?:\d+\.?\d*|\.\d+)", str(value or "").replace(",", "."))
    return float(match.group(0)) if match else None


def gcode_number(value: float, digits: int = 3) -> str:
    return f"{value:.{digits}f}".rstrip("0").rstrip(".") or "0"


def gcode_label(value: Any, limit: int = 40) -> str:
    label = re.sub(r"[^A-Za-z0-9_.-]", "_", str(value or ""))[:limit]
    return label or "item"


class PathPreview(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.points: list[PointData] = []
        self.pads: list[GerberPad] = []
        self.segments: list[DispenseSegment] = []
        self.setMinimumSize(420, 280)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

    def set_points(self, points: list[PointData]) -> None:
        self.set_geometry(points, [], [])

    def set_geometry(self, points: list[PointData],
                     pads: list[GerberPad],
                     segments: list[DispenseSegment]) -> None:
        self.points = points
        self.pads = pads
        self.segments = segments
        self.update()

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QColor("#ffffff"))
        painter.setPen(QPen(QColor("#e6eaeb"), 1))
        for x in range(0, self.width(), 24):
            painter.drawLine(x, 0, x, self.height())
        for y in range(0, self.height(), 24):
            painter.drawLine(0, y, self.width(), y)
        if not self.points and not self.pads:
            painter.setPen(QColor("#68777d"))
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "没有选中的点胶对象")
            return

        xs = [p.x for p in self.points]
        ys = [p.y for p in self.points]
        for pad in self.pads:
            xs.extend(x for x, _ in pad.vertices)
            ys.extend(y for _, y in pad.vertices)
        for segment in self.segments:
            xs.extend([segment.start[0], segment.end[0]])
            ys.extend([segment.start[1], segment.end[1]])
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        range_x, range_y = max(1.0, max_x - min_x), max(1.0, max_y - min_y)
        drawing = self.rect().adjusted(28, 22, -28, -22)
        scale = min(drawing.width() / range_x, drawing.height() / range_y)
        board_w, board_h = range_x * scale, range_y * scale
        left = drawing.center().x() - board_w / 2
        top = drawing.center().y() - board_h / 2

        def map_xy(x: float, y: float) -> QPointF:
            return QPointF(left + (x - min_x) * scale, top + (max_y - y) * scale)

        painter.setPen(QPen(QColor("#dbe2e3"), 1, Qt.PenStyle.DashLine))
        painter.setBrush(QColor(246, 249, 248, 190))
        painter.drawRect(QRectF(left, top, board_w, board_h))
        if self.pads:
            painter.setPen(QPen(QColor("#087f5b"), 1))
            painter.setBrush(QColor(8, 127, 91, 46))
            for pad in self.pads:
                painter.drawPolygon(QPolygonF([map_xy(x, y) for x, y in pad.vertices]))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.setPen(QPen(QColor("#d94841"), 1.35))
            for segment in self.segments:
                painter.drawLine(map_xy(*segment.start), map_xy(*segment.end))
        if self.points:
            path = QPainterPath(map_xy(self.points[0].x, self.points[0].y))
            for point in self.points[1:]:
                path.lineTo(map_xy(point.x, point.y))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.setPen(QPen(QColor(224, 122, 34, 145), 1.15))
            painter.drawPath(path)
            for index, point in enumerate(self.points):
                painter.setPen(QPen(QColor("#ffffff"), 1))
                painter.setBrush(QColor("#d94841" if index == 0 else "#087f5b"))
                painter.drawEllipse(map_xy(point.x, point.y), 3.8, 3.8)


class MainWindow(QMainWindow):
    def __init__(self, initial_file: Path | None = None, initial_gerber: Path | None = None) -> None:
        super().__init__()
        self.sheets: list[SheetData] = []
        self.rows: list[dict[str, Any]] = []
        self.selected_ids: set[int] = set()
        self.gerber_pads: list[GerberPad] = []
        self.selected_pad_ids: set[int] = set()
        self.current_gerber: Path | None = None
        self.gerber_warnings: list[str] = []
        self.current_file: Path | None = None
        self._updating_table = False
        self.devices: list[WirelessDevice] = []
        self.connected_device: WirelessDevice | None = None
        self.saved_device: WirelessDevice | None = self.load_saved_device()
        self.bridge_socket: socket.socket | None = None
        self._connecting = False
        self._transfer_busy = False
        self._terminal_busy = False
        self._prev_ssid = ""
        self.setWindowTitle("Banux PCB 锡膏路径生成器")
        self.resize(1440, 860)
        self.setMinimumSize(1050, 680)
        self._build_ui()
        self._connect_signals()
        self._apply_style()
        # 启动后如果已经有 saved_device，先预填 UI 让用户感知"已记住设备"
        if self.saved_device and self.saved_device.ip:
            self.device_name_label.setText(self.saved_device.name or DEFAULT_DEVICE_NAME)
            suffix = " AP" if self.saved_device.ap else ""
            hint = (
                f"{self.saved_device.name or DEFAULT_DEVICE_NAME}  "
                f"{self.saved_device.ip}:{self.saved_device.bridge_port}{suffix}"
                f"  WiFi: {self.saved_device.wifi or '—'}"
                "   （自动重连中…）"
            )
            self.device_info_label.setText(hint)
            self.set_device_state(
                "offline", f"记住上次设备，重连中 {self.saved_device.ip}:{self.saved_device.bridge_port}")
        target = initial_file
        if target:
            QTimer.singleShot(0, lambda: self.import_workbook(target, show_error=False))
        if initial_gerber:
            QTimer.singleShot(50, lambda: self.import_gerber(initial_gerber, show_error=False))
        self.device_timer = QTimer(self)
        self.device_timer.setInterval(3000)
        self.device_timer.timeout.connect(self.poll_device_status)
        self.device_timer.start()
        # 延时 900ms 再做首次探测（给 Windows 多网卡/WiFi 栈就绪留时间）
        QTimer.singleShot(900, self.poll_device_status)

    def _build_ui(self) -> None:
        toolbar = QToolBar("主工具栏")
        toolbar.setMovable(False)
        toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.addToolBar(toolbar)
        title = QLabel("PCB 锡膏路径生成器")
        title.setObjectName("appTitle")
        toolbar.addWidget(title)
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(spacer)
        self.import_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton), "导入 XLSX", self)
        self.import_gerber_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogDetailedView), "导入 Gerber", self)
        self.export_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton), "导出 G-code", self)
        self.start_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay), "开始", self)
        self.settings_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogDetailedView), "无线设置", self)
        self.terminal_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon), "命令行", self)
        toolbar.addAction(self.import_gerber_action)
        toolbar.addAction(self.export_action)
        toolbar.addAction(self.start_action)
        toolbar.addAction(self.settings_action)
        toolbar.addAction(self.terminal_action)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.addWidget(self._build_settings())
        splitter.addWidget(self._build_center())
        splitter.addWidget(self._build_points())
        splitter.setSizes([290, 760, 390])
        splitter.setStretchFactor(1, 1)
        self.prepare_page = splitter
        self.pages = QStackedWidget()
        self.pages.addWidget(self.prepare_page)
        self.start_page = self._build_start_page()
        self.pages.addWidget(self.start_page)
        self.setCentralWidget(self.pages)
        self.setStatusBar(QStatusBar())
        self.statusBar().showMessage("请选择 Gerber Paste 文件")
        self.update_start_enabled()

    def _spin(self, value: float, minimum: float = -100000.0,
              maximum: float = 100000.0, step: float = 0.1,
              decimals: int = 3) -> QDoubleSpinBox:
        box = QDoubleSpinBox()
        box.setRange(minimum, maximum)
        box.setDecimals(decimals)
        box.setSingleStep(step)
        box.setValue(value)
        box.setKeyboardTracking(False)
        return box

    def _build_settings(self) -> QWidget:
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(12, 12, 12, 16)
        layout.setSpacing(12)
        coordinate_group = QGroupBox("坐标设置")
        form = QFormLayout(coordinate_group)
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["坐标表点胶", "Gerber 焊盘铺膏"])
        self.sheet_combo = QComboBox()
        self.source_combo = QComboBox()
        self.source_combo.addItems(["Mid X / Mid Y", "Ref X / Ref Y", "Pad X / Pad Y"])
        self.origin_combo = QComboBox()
        self.origin_combo.addItems(["左下角归零", "保留原坐标", "指定工件原点"])
        self.offset_x, self.offset_y = self._spin(0), self._spin(0)
        self.origin_x, self.origin_y = self._spin(0), self._spin(0)
        self.flip_x, self.flip_y = QCheckBox("X 轴镜像"), QCheckBox("Y 轴镜像")
        mirror_row = QWidget()
        mirror_layout = QHBoxLayout(mirror_row)
        mirror_layout.setContentsMargins(0, 0, 0, 0)
        mirror_layout.addWidget(self.flip_x)
        mirror_layout.addWidget(self.flip_y)
        form.addRow("路径来源", self.mode_combo)
        form.addRow("工作表", self.sheet_combo)
        form.addRow("坐标列", self.source_combo)
        form.addRow("原点", self.origin_combo)
        form.addRow("X 偏移 (mm)", self.offset_x)
        form.addRow("Y 偏移 (mm)", self.offset_y)
        form.addRow("原点 X (mm)", self.origin_x)
        form.addRow("原点 Y (mm)", self.origin_y)
        form.addRow("镜像", mirror_row)

        motion_group = QGroupBox("运动参数")
        motion = QFormLayout(motion_group)
        self.safe_z = self._spin(5, 0, 1000, 0.1)
        self.paste_z = self._spin(0.2, -1000, 1000, 0.05)
        self.travel_feed = self._spin(1200, 1, 100000, 50, 0)
        self.z_feed = self._spin(300, 1, 100000, 10, 0)
        self.extrude = self._spin(0.5, 0, 1000, 0.01)
        self.retract = self._spin(0.05, 0, 1000, 0.01)
        self.e_feed = self._spin(60, 1, 100000, 5, 0)
        self.nozzle_diameter = self._spin(0.35, 0.05, 10, 0.05)
        self.line_spacing = self._spin(0.30, 0.05, 10, 0.05)
        self.edge_inset = self._spin(0.08, 0, 10, 0.02)
        self.e_per_mm = self._spin(0.08, 0, 1000, 0.005, 4)
        self.min_stroke_length = self._spin(0.25, 0, 1000, 0.05)
        self.order_combo = QComboBox()
        self.order_combo.addItems(["坐标表原顺序", "最近点优先"])
        for label, widget in [("安全 Z (mm)", self.safe_z), ("点胶 Z (mm)", self.paste_z),
                              ("XY 速度", self.travel_feed), ("Z 速度", self.z_feed),
                              ("单点挤出 E", self.extrude), ("回抽 E", self.retract),
                              ("E 轴速度", self.e_feed), ("喷嘴直径 (mm)", self.nozzle_diameter),
                              ("铺膏线距 (mm)", self.line_spacing), ("焊盘内缩 (mm)", self.edge_inset),
                              ("铺膏 E/mm", self.e_per_mm), ("最短线段 (mm)", self.min_stroke_length),
                              ("路径顺序", self.order_combo)]:
            motion.addRow(label, widget)
        layout.addWidget(coordinate_group)
        layout.addWidget(motion_group)
        layout.addStretch()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(content)
        scroll.setMinimumWidth(270)
        return scroll

    def _build_center(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(9)
        title_row = QHBoxLayout()
        title = QLabel("路径预览")
        title.setObjectName("sectionTitle")
        self.board_size_label = QLabel("--")
        self.board_size_label.setObjectName("muted")
        title_row.addWidget(title)
        title_row.addStretch()
        title_row.addWidget(self.board_size_label)
        layout.addLayout(title_row)
        self.preview = PathPreview()
        self.preview.setObjectName("preview")
        layout.addWidget(self.preview, 3)
        stats = QFrame()
        stats.setObjectName("stats")
        stats_layout = QGridLayout(stats)
        stats_layout.setContentsMargins(10, 8, 10, 8)
        self.point_count, self.distance_label = QLabel("0"), QLabel("0 mm")
        self.time_label, self.line_count = QLabel("0 s"), QLabel("0")
        for column, (label, value) in enumerate([
            ("点胶点", self.point_count), ("空移距离", self.distance_label),
            ("预计耗时", self.time_label), ("G-code 行数", self.line_count),
        ]):
            caption = QLabel(label)
            caption.setObjectName("muted")
            value.setObjectName("statValue")
            stats_layout.addWidget(caption, 0, column)
            stats_layout.addWidget(value, 1, column)
        layout.addWidget(stats)
        code_header = QHBoxLayout()
        code_title = QLabel("G-code 预览")
        code_title.setObjectName("sectionTitle")
        self.copy_button = QPushButton("复制")
        self.go_start_button = QPushButton("开始")
        code_header.addWidget(code_title)
        code_header.addStretch()
        code_header.addWidget(self.copy_button)
        code_header.addWidget(self.go_start_button)
        layout.addLayout(code_header)
        self.gcode_preview = QPlainTextEdit()
        self.gcode_preview.setReadOnly(True)
        self.gcode_preview.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        self.gcode_preview.setFont(QFont("Consolas", 9))
        layout.addWidget(self.gcode_preview, 2)
        return panel

    def _build_start_page(self) -> QWidget:
        page = QWidget()
        page_layout = QHBoxLayout(page)
        page_layout.setContentsMargins(18, 16, 18, 16)
        page_layout.setSpacing(12)
        left_panel = QWidget()
        layout = QVBoxLayout(left_panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        connection = QGroupBox("设备")
        grid = QGridLayout(connection)
        self.device_combo = QComboBox()
        self.device_combo.setVisible(False)
        self.device_status_label = QLabel("离线")
        self.device_status_label.setObjectName("deviceOffline")
        self.device_name_label = QLabel(self.saved_device.name if self.saved_device else DEFAULT_DEVICE_NAME)
        self.device_name_label.setObjectName("statValue")
        self.device_info_label = QLabel("正在查找设备")
        self.device_info_label.setObjectName("muted")
        self.first_setup_button = QPushButton("首次设置")
        grid.addWidget(self.device_status_label, 0, 0)
        grid.addWidget(self.device_name_label, 0, 1, 1, 3)
        grid.addWidget(self.first_setup_button, 0, 4)
        grid.addWidget(self.device_info_label, 1, 0, 1, 5)

        transfer = QGroupBox("G-code 传输与执行")
        transfer_grid = QGridLayout(transfer)
        self.storage_combo = QComboBox()
        self.storage_combo.addItems(["/flash", "/sd"])
        self.remote_path_edit = QLineEdit("/flash/pcb_solder_paste.gcode")
        self.upload_button = QPushButton("传到设备")
        self.execute_button = QPushButton("执行")
        self.transfer_progress = QLabel("0%")
        self.transfer_progress.setObjectName("statValue")
        transfer_grid.addWidget(QLabel("存储"), 0, 0)
        transfer_grid.addWidget(self.storage_combo, 0, 1)
        transfer_grid.addWidget(QLabel("路径"), 0, 2)
        transfer_grid.addWidget(self.remote_path_edit, 0, 3, 1, 3)
        transfer_grid.addWidget(self.upload_button, 0, 6)
        transfer_grid.addWidget(self.execute_button, 0, 7)
        transfer_grid.addWidget(QLabel("进度"), 1, 0)
        transfer_grid.addWidget(self.transfer_progress, 1, 1, 1, 7)

        self.connection_log = QPlainTextEdit()
        self.connection_log.setReadOnly(True)
        self.connection_log.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        self.connection_log.setFont(QFont("Consolas", 9))
        self.connection_log.setVisible(False)

        layout.addWidget(connection)
        layout.addWidget(transfer)
        layout.addWidget(self.connection_log)
        layout.addStretch(1)
        page_layout.addWidget(left_panel, 1)
        page_layout.addWidget(self._build_manual_console())
        return page

    def _build_manual_console(self) -> QWidget:
        panel = QGroupBox("手动操作台")
        panel.setMinimumWidth(280)
        layout = QVBoxLayout(panel)
        layout.setSpacing(10)

        motion = QFormLayout()
        self.jog_step = self._spin(1.0, 0.01, 100.0, 0.1)
        self.jog_feed = self._spin(600.0, 1.0, 10000.0, 50.0, 0)
        self.extruder_step = self._spin(0.2, 0.001, 100.0, 0.01, 3)
        motion.addRow("步距 (mm)", self.jog_step)
        motion.addRow("速度", self.jog_feed)
        motion.addRow("挤出量 E", self.extruder_step)
        layout.addLayout(motion)

        xyz_grid = QGridLayout()
        self.y_plus_button = QPushButton("Y+")
        self.y_minus_button = QPushButton("Y-")
        self.x_minus_button = QPushButton("X-")
        self.x_plus_button = QPushButton("X+")
        self.z_plus_button = QPushButton("Z+")
        self.z_minus_button = QPushButton("Z-")
        xyz_grid.addWidget(self.y_plus_button, 0, 1)
        xyz_grid.addWidget(self.x_minus_button, 1, 0)
        xyz_grid.addWidget(self.x_plus_button, 1, 2)
        xyz_grid.addWidget(self.y_minus_button, 2, 1)
        xyz_grid.addWidget(self.z_plus_button, 0, 3)
        xyz_grid.addWidget(self.z_minus_button, 2, 3)
        layout.addLayout(xyz_grid)

        extruder_row = QHBoxLayout()
        self.e_minus_button = QPushButton("E-")
        self.e_plus_button = QPushButton("E+")
        extruder_row.addWidget(self.e_minus_button)
        extruder_row.addWidget(self.e_plus_button)
        layout.addLayout(extruder_row)

        utility_row = QHBoxLayout()
        self.motor_on_button = QPushButton("使能")
        self.motor_off_button = QPushButton("释放")
        self.status_button = QPushButton("状态")
        self.manual_buttons = [
            self.y_plus_button, self.y_minus_button, self.x_minus_button,
            self.x_plus_button, self.z_plus_button, self.z_minus_button,
            self.e_minus_button, self.e_plus_button, self.motor_on_button,
            self.motor_off_button, self.status_button,
        ]
        utility_row.addWidget(self.motor_on_button)
        utility_row.addWidget(self.motor_off_button)
        utility_row.addWidget(self.status_button)
        layout.addLayout(utility_row)
        layout.addStretch()
        return panel

    def _build_points(self) -> QWidget:
        panel = QWidget()
        panel.setMinimumWidth(340)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(10, 12, 10, 10)
        layout.setSpacing(8)
        header = QHBoxLayout()
        title = QLabel("点胶对象")
        title.setObjectName("sectionTitle")
        self.summary_label = QLabel("0 / 0")
        self.summary_label.setObjectName("muted")
        header.addWidget(title)
        header.addStretch()
        header.addWidget(self.summary_label)
        layout.addLayout(header)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索位号、封装、器件或焊盘")
        layout.addWidget(self.search_edit)
        filters = QHBoxLayout()
        self.layer_combo = QComboBox()
        self.layer_combo.addItems(["全部层", "顶层", "底层"])
        self.smd_only = QCheckBox("仅 SMD")
        self.smd_only.setChecked(True)
        filters.addWidget(self.layer_combo)
        filters.addWidget(self.smd_only)
        filters.addStretch()
        layout.addLayout(filters)
        selection = QHBoxLayout()
        self.select_button, self.clear_button = QPushButton("选择可见"), QPushButton("清除可见")
        selection.addWidget(self.select_button)
        selection.addWidget(self.clear_button)
        selection.addStretch()
        layout.addLayout(selection)
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["选", "对象", "坐标 (mm)", "信息"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setColumnWidth(0, 42)
        self.table.setColumnWidth(1, 70)
        self.table.setColumnWidth(2, 112)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)
        return panel

    def _connect_signals(self) -> None:
        self.import_action.triggered.connect(self.choose_workbook)
        self.import_gerber_action.triggered.connect(self.choose_gerber)
        self.export_action.triggered.connect(self.export_gcode)
        self.start_action.triggered.connect(self.show_start_page)
        self.settings_action.triggered.connect(self.show_wireless_settings)
        self.terminal_action.triggered.connect(self.show_bridge_terminal)
        self.sheet_combo.currentIndexChanged.connect(self.sheet_changed)
        for widget in [self.mode_combo, self.source_combo, self.origin_combo, self.order_combo, self.layer_combo]:
            widget.currentIndexChanged.connect(self.refresh)
        for widget in [self.flip_x, self.flip_y, self.smd_only]:
            widget.toggled.connect(self.refresh)
        for spin in [self.offset_x, self.offset_y, self.origin_x, self.origin_y,
                     self.safe_z, self.paste_z, self.travel_feed, self.z_feed,
                     self.extrude, self.retract, self.e_feed, self.nozzle_diameter,
                     self.line_spacing, self.edge_inset, self.e_per_mm, self.min_stroke_length]:
            spin.valueChanged.connect(self.refresh)
        self.search_edit.textChanged.connect(self.refresh_table)
        self.table.itemChanged.connect(self.table_item_changed)
        self.select_button.clicked.connect(lambda: self.select_visible(True))
        self.clear_button.clicked.connect(lambda: self.select_visible(False))
        self.copy_button.clicked.connect(self.copy_gcode)
        self.go_start_button.clicked.connect(self.show_start_page)
        self.first_setup_button.clicked.connect(self.start_first_setup)
        self.storage_combo.currentTextChanged.connect(self.update_remote_path_root)
        self.upload_button.clicked.connect(self.upload_gcode_to_device)
        self.execute_button.clicked.connect(self.execute_remote_gcode)
        self.x_minus_button.clicked.connect(lambda _checked=False: self.jog_axis("X", -1.0))
        self.x_plus_button.clicked.connect(lambda _checked=False: self.jog_axis("X", 1.0))
        self.y_minus_button.clicked.connect(lambda _checked=False: self.jog_axis("Y", -1.0))
        self.y_plus_button.clicked.connect(lambda _checked=False: self.jog_axis("Y", 1.0))
        self.z_minus_button.clicked.connect(lambda _checked=False: self.jog_axis("Z", -1.0))
        self.z_plus_button.clicked.connect(lambda _checked=False: self.jog_axis("Z", 1.0))
        self.e_minus_button.clicked.connect(lambda _checked=False: self.jog_axis("E", -1.0))
        self.e_plus_button.clicked.connect(lambda _checked=False: self.jog_axis("E", 1.0))
        self.motor_on_button.clicked.connect(lambda _checked=False: self.send_manual_command("gcode M17"))
        self.motor_off_button.clicked.connect(lambda _checked=False: self.send_manual_command("gcode M84"))
        self.status_button.clicked.connect(lambda _checked=False: self.send_manual_command("gcode -s"))

    def _apply_style(self) -> None:
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #f5f7f7; color: #1c282d; font-family: "Microsoft YaHei UI"; font-size: 12px; }
            QToolBar { min-height: 52px; padding: 4px 12px; spacing: 8px; background: #202b30; border-bottom: 3px solid #087f5b; }
            QToolBar QToolButton { min-height: 30px; padding: 3px 10px; color: white; background: #2e3b40; border: 1px solid #5b696e; border-radius: 4px; }
            QToolBar QToolButton:hover { background: #3a494f; }
            QToolBar QWidget { background: transparent; }
            QLabel#appTitle { min-width: 220px; color: white; background: transparent; font-size: 18px; font-weight: 700; }
            QLabel#sectionTitle { font-size: 14px; font-weight: 700; }
            QLabel#muted { color: #68777d; }
            QLabel#statValue { font-size: 16px; font-weight: 700; }
            QLabel#deviceOnline { min-width: 56px; padding: 5px 9px; color: white; background: #087f5b; border-radius: 4px; font-weight: 700; }
            QLabel#deviceOffline { min-width: 56px; padding: 5px 9px; color: white; background: #8b1e2d; border-radius: 4px; font-weight: 700; }
            QLabel#deviceSetup { min-width: 56px; padding: 5px 9px; color: #1c282d; background: #ffd43b; border-radius: 4px; font-weight: 700; }
            QGroupBox { margin-top: 9px; padding-top: 11px; font-weight: 700; border: 1px solid #ccd4d6; border-radius: 5px; background: white; }
            QGroupBox::title { subcontrol-origin: margin; left: 9px; padding: 0 4px; }
            QLineEdit, QComboBox, QDoubleSpinBox { min-height: 28px; padding: 0 6px; background: white; border: 1px solid #bdc7ca; border-radius: 4px; }
            QLineEdit:focus, QComboBox:focus, QDoubleSpinBox:focus { border: 1px solid #087f5b; }
            QPushButton { min-height: 29px; padding: 0 11px; background: white; border: 1px solid #aeb9bc; border-radius: 4px; }
            QPushButton:hover { background: #edf7f3; border-color: #087f5b; }
            QWidget#preview { border: 1px solid #c8d0d2; }
            QFrame#stats { background: white; border: 1px solid #d5dcde; }
            QPlainTextEdit { color: #bdebdc; background: #182126; border: 1px solid #34444a; }
            QTableWidget { background: white; alternate-background-color: #f7f9f9; border: 1px solid #ccd4d6; gridline-color: #e5e9ea; }
            QHeaderView::section { padding: 6px; color: #526168; background: #eef2f2; border: 0; border-bottom: 1px solid #cbd3d5; }
            QSplitter::handle { background: #d5dcde; width: 1px; }
        """)

    def choose_workbook(self) -> None:
        initial = str(self.current_file.parent if self.current_file else ROOT.parent)
        file_name, _ = QFileDialog.getOpenFileName(self, "选择 PCB 坐标表", initial, "Excel 工作簿 (*.xlsx)")
        if file_name:
            self.import_workbook(Path(file_name))

    def choose_gerber(self) -> None:
        initial = str(self.current_gerber.parent if self.current_gerber else ROOT.parent)
        file_name, _ = QFileDialog.getOpenFileName(
            self, "选择 Gerber Paste 文件", initial,
            "Gerber Paste 或压缩包 (*.gtp *.gbp *.zip);;Gerber 文件 (*.gbr *.ger *.pho *.art);;所有文件 (*.*)")
        if file_name:
            self.import_gerber(Path(file_name))

    def import_workbook(self, path: Path, show_error: bool = True) -> bool:
        try:
            sheets = load_workbook(path)
            if not any({"Mid X", "Mid Y"}.issubset(set(sheet.headers)) for sheet in sheets):
                raise ValueError("找不到 Mid X / Mid Y 坐标列")
        except Exception as exc:
            if show_error:
                QMessageBox.critical(self, "导入失败", f"无法读取坐标表：\n{exc}")
            self.statusBar().showMessage(f"导入失败：{exc}")
            return False
        self.sheets, self.current_file = sheets, path
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for sheet in sheets:
            self.sheet_combo.addItem(f"{sheet.name} ({len(sheet.rows)})")
        self.sheet_combo.blockSignals(False)
        self.sheet_combo.setCurrentIndex(0)
        self.sheet_changed(0)
        self.statusBar().showMessage(f"已导入 {path.name}，共 {sum(len(s.rows) for s in sheets)} 条记录")
        return True

    def import_gerber(self, path: Path, show_error: bool = True) -> bool:
        try:
            result = parse_gerber_file(path)
            if not result.pads:
                raise ValueError("没有解析到焊盘。请优先选择 Top Paste(.GTP) 或 Bottom Paste(.GBP) 层。")
        except Exception as exc:
            if show_error:
                QMessageBox.critical(self, "导入失败", f"无法读取 Gerber：\n{exc}")
            self.statusBar().showMessage(f"Gerber 导入失败：{exc}")
            return False
        self.gerber_pads = result.pads
        self.selected_pad_ids = {pad.pad_id for pad in result.pads}
        self.current_gerber = path
        self.gerber_warnings = result.warnings
        self.mode_combo.setCurrentIndex(1)
        self.refresh()
        warning = f"，{len(result.warnings)} 条警告" if result.warnings else ""
        self.statusBar().showMessage(
            f"已导入 {path.name}，解析 {len(result.pads)} 个焊盘，单位 {result.unit}，格式 {result.coordinate_format}{warning}")
        return True

    def sheet_changed(self, index: int) -> None:
        self.rows = self.sheets[index].rows if 0 <= index < len(self.sheets) else []
        self.selected_ids = {i for i, row in enumerate(self.rows)
                             if str(row.get("SMD", "")).lower() == "yes"}
        self.refresh()

    def transformed_points(self) -> list[PointData]:
        prefix = ("Mid", "Ref", "Pad")[self.source_combo.currentIndex()]
        points: list[PointData] = []
        for row_id, row in enumerate(self.rows):
            x, y = parse_number(row.get(f"{prefix} X")), parse_number(row.get(f"{prefix} Y"))
            if x is not None and y is not None:
                points.append(PointData(row_id, row, x, y))
        if not points:
            return []
        mode = self.origin_combo.currentIndex()
        flip_x, flip_y = self.flip_x.isChecked(), self.flip_y.isChecked()
        if mode == 0:
            min_x, max_x = min(p.x for p in points), max(p.x for p in points)
            min_y, max_y = min(p.y for p in points), max(p.y for p in points)
            for point in points:
                point.x = (max_x - point.x if flip_x else point.x - min_x) + self.offset_x.value()
                point.y = (max_y - point.y if flip_y else point.y - min_y) + self.offset_y.value()
        else:
            origin_x = self.origin_x.value() if mode == 2 else 0.0
            origin_y = self.origin_y.value() if mode == 2 else 0.0
            for point in points:
                point.x = (point.x - origin_x) * (-1 if flip_x else 1) + self.offset_x.value()
                point.y = (point.y - origin_y) * (-1 if flip_y else 1) + self.offset_y.value()
        return points

    def transformed_pads(self) -> list[GerberPad]:
        if not self.gerber_pads:
            return []
        all_x = [x for pad in self.gerber_pads for x, _ in pad.vertices]
        all_y = [y for pad in self.gerber_pads for _, y in pad.vertices]
        min_x, max_x = min(all_x), max(all_x)
        min_y, max_y = min(all_y), max(all_y)
        mode = self.origin_combo.currentIndex()
        flip_x, flip_y = self.flip_x.isChecked(), self.flip_y.isChecked()
        origin_x = self.origin_x.value() if mode == 2 else 0.0
        origin_y = self.origin_y.value() if mode == 2 else 0.0

        def mapped(x: float, y: float) -> tuple[float, float]:
            if mode == 0:
                tx = max_x - x if flip_x else x - min_x
                ty = max_y - y if flip_y else y - min_y
            else:
                tx = (x - origin_x) * (-1 if flip_x else 1)
                ty = (y - origin_y) * (-1 if flip_y else 1)
            return tx + self.offset_x.value(), ty + self.offset_y.value()

        pads: list[GerberPad] = []
        for pad in self.gerber_pads:
            vertices = [mapped(x, y) for x, y in pad.vertices]
            xs = [x for x, _ in vertices]
            ys = [y for _, y in vertices]
            cx, cy = mapped(pad.x, pad.y)
            pads.append(GerberPad(pad.pad_id, cx, cy, max(xs) - min(xs), max(ys) - min(ys),
                                  pad.shape, vertices, pad.source_layer, pad.aperture,
                                  pad.rotation, pad.label))
        return pads

    def visible_points(self) -> list[PointData]:
        query = self.search_edit.text().strip().lower()
        layer = (None, "T", "B")[self.layer_combo.currentIndex()]
        result = []
        for point in self.transformed_points():
            raw = point.raw
            if self.smd_only.isChecked() and str(raw.get("SMD", "")).lower() != "yes":
                continue
            if layer and str(raw.get("Layer", "")).upper() != layer:
                continue
            text = " ".join(str(raw.get(k, "")) for k in ("Designator", "Footprint", "Device", "Comment")).lower()
            if not query or query in text:
                result.append(point)
        return result

    def visible_pads(self) -> list[GerberPad]:
        query = self.search_edit.text().strip().lower()
        layer_index = self.layer_combo.currentIndex()
        result: list[GerberPad] = []
        for pad in self.transformed_pads():
            layer = pad.source_layer.lower()
            if layer_index == 1 and "bottom" in layer:
                continue
            if layer_index == 2 and "top" in layer:
                continue
            text = f"{pad.label} {pad.shape} {pad.aperture} {pad.source_layer} {pad.width:.3f} {pad.height:.3f}".lower()
            if not query or query in text:
                result.append(pad)
        return result

    def ordered_selected_points(self) -> list[PointData]:
        points = [point for point in self.transformed_points() if point.row_id in self.selected_ids]
        if self.order_combo.currentIndex() == 0 or len(points) < 2:
            return points
        pending, ordered = points[:], []
        ordered.append(pending.pop(0))
        while pending:
            previous = ordered[-1]
            best = min(range(len(pending)), key=lambda i: math.hypot(pending[i].x - previous.x,
                                                                      pending[i].y - previous.y))
            ordered.append(pending.pop(best))
        return ordered

    def ordered_selected_pads(self) -> list[GerberPad]:
        pads = [pad for pad in self.transformed_pads() if pad.pad_id in self.selected_pad_ids]
        if self.order_combo.currentIndex() == 0 or len(pads) < 2:
            return pads
        pending, ordered = pads[:], []
        ordered.append(pending.pop(0))
        while pending:
            previous = ordered[-1]
            best = min(range(len(pending)), key=lambda i: math.hypot(pending[i].x - previous.x,
                                                                      pending[i].y - previous.y))
            ordered.append(pending.pop(best))
        return ordered

    def pad_dispense_segments(self, pads: list[GerberPad] | None = None) -> list[DispenseSegment]:
        pads = self.ordered_selected_pads() if pads is None else pads
        segments: list[DispenseSegment] = []
        for pad in pads:
            segments.extend(self._segments_for_pad(pad))
        return segments

    def _segments_for_pad(self, pad: GerberPad) -> list[DispenseSegment]:
        spacing = max(0.05, self.line_spacing.value())
        inset = max(0.0, self.edge_inset.value())
        min_length = max(0.0, self.min_stroke_length.value())
        xs = [x for x, _ in pad.vertices]
        ys = [y for _, y in pad.vertices]
        min_x, max_x = min(xs) + inset, max(xs) - inset
        min_y, max_y = min(ys) + inset, max(ys) - inset
        if max_x <= min_x or max_y <= min_y or min(pad.width, pad.height) < self.nozzle_diameter.value() * 1.15:
            return [DispenseSegment(pad.pad_id, pad.label, (pad.x, pad.y), (pad.x, pad.y))]

        height = max_y - min_y
        line_count = max(1, int(math.floor(height / spacing)) + 1)
        if line_count == 1:
            y_values = [(min_y + max_y) / 2.0]
        else:
            margin = max(0.0, (height - spacing * (line_count - 1)) / 2.0)
            y_values = [min_y + margin + i * spacing for i in range(line_count)]
        result: list[DispenseSegment] = []
        reverse = False
        for y in y_values:
            spans = self._polygon_spans_at_y(pad.vertices, y)
            for left, right in spans:
                left += inset
                right -= inset
                if right - left < min_length:
                    continue
                start, end = ((right, y), (left, y)) if reverse else ((left, y), (right, y))
                result.append(DispenseSegment(pad.pad_id, pad.label, start, end))
                reverse = not reverse
        if not result:
            result.append(DispenseSegment(pad.pad_id, pad.label, (pad.x, pad.y), (pad.x, pad.y)))
        return result

    @staticmethod
    def _polygon_spans_at_y(vertices: list[tuple[float, float]], y: float) -> list[tuple[float, float]]:
        hits: list[float] = []
        if len(vertices) < 3:
            return []
        for index, (x1, y1) in enumerate(vertices):
            x2, y2 = vertices[(index + 1) % len(vertices)]
            if abs(y1 - y2) < 1e-9:
                continue
            if (y1 <= y < y2) or (y2 <= y < y1):
                t = (y - y1) / (y2 - y1)
                hits.append(x1 + t * (x2 - x1))
        hits.sort()
        spans: list[tuple[float, float]] = []
        for i in range(0, len(hits) - 1, 2):
            if hits[i + 1] > hits[i]:
                spans.append((hits[i], hits[i + 1]))
        return spans

    def refresh(self, *_args) -> None:
        custom = self.origin_combo.currentIndex() == 2
        self.origin_x.setEnabled(custom)
        self.origin_y.setEnabled(custom)
        gerber_mode = self.mode_combo.currentIndex() == 1
        self.sheet_combo.setEnabled(not gerber_mode)
        self.source_combo.setEnabled(not gerber_mode)
        self.smd_only.setEnabled(not gerber_mode)
        self.refresh_table()
        points = [] if gerber_mode else self.ordered_selected_points()
        pads = self.ordered_selected_pads() if gerber_mode else []
        segments = self.pad_dispense_segments(pads) if gerber_mode else []
        self.preview.set_geometry(points, pads, segments)
        self._update_board_size(points, pads)
        gcode = self.generate_gcode()
        self.gcode_preview.setPlainText(gcode)
        distance, seconds = self.path_stats(points, segments)
        self.point_count.setText(str(len(pads) if gerber_mode else len(points)))
        self.distance_label.setText(f"{distance:.1f} mm")
        self.time_label.setText(f"{seconds:.0f} s" if seconds < 60 else f"{seconds / 60:.1f} min")
        self.line_count.setText(str(len(gcode.rstrip().splitlines())))
        self.update_start_enabled()

    def _update_board_size(self, points: list[PointData], pads: list[GerberPad]) -> None:
        xs = [p.x for p in points]
        ys = [p.y for p in points]
        for pad in pads:
            xs.extend(x for x, _ in pad.vertices)
            ys.extend(y for _, y in pad.vertices)
        if xs and ys:
            self.board_size_label.setText(f"{max(xs) - min(xs):.2f} x {max(ys) - min(ys):.2f} mm")
        else:
            self.board_size_label.setText("--")

    def refresh_table(self, *_args) -> None:
        gerber_mode = self.mode_combo.currentIndex() == 1
        visible_points = [] if gerber_mode else self.visible_points()
        visible_pads = self.visible_pads() if gerber_mode else []
        self._updating_table = True
        self.table.setRowCount(len(visible_pads) if gerber_mode else len(visible_points))
        if gerber_mode:
            for table_row, pad in enumerate(visible_pads):
                check = QTableWidgetItem()
                check.setData(Qt.ItemDataRole.UserRole, pad.pad_id)
                check.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable)
                check.setCheckState(Qt.CheckState.Checked if pad.pad_id in self.selected_pad_ids else Qt.CheckState.Unchecked)
                name = QTableWidgetItem(pad.label)
                name.setToolTip(f"{pad.source_layer}  {pad.aperture}")
                coordinate = QTableWidgetItem(f"X {gcode_number(pad.x)}\nY {gcode_number(pad.y)}")
                info = QTableWidgetItem(f"{pad.shape} {pad.width:.3f} x {pad.height:.3f}")
                info.setToolTip(f"shape={pad.shape}, aperture={pad.aperture}, layer={pad.source_layer}")
                for column, item in enumerate((check, name, coordinate, info)):
                    self.table.setItem(table_row, column, item)
                self.table.setRowHeight(table_row, 43)
            selected_count = sum(pad.pad_id in self.selected_pad_ids for pad in visible_pads)
            self.summary_label.setText(f"{selected_count} / {len(visible_pads)} 可见")
            self._updating_table = False
            return
        for table_row, point in enumerate(visible_points):
            check = QTableWidgetItem()
            check.setData(Qt.ItemDataRole.UserRole, point.row_id)
            check.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable)
            check.setCheckState(Qt.CheckState.Checked if point.row_id in self.selected_ids else Qt.CheckState.Unchecked)
            raw = point.raw
            designation = QTableWidgetItem(str(raw.get("Designator", "-")))
            designation.setToolTip(f"层: {raw.get('Layer', '-')}  旋转: {raw.get('Rotation', 0)}°")
            coordinate = QTableWidgetItem(f"X {gcode_number(point.x)}\nY {gcode_number(point.y)}")
            footprint = QTableWidgetItem(str(raw.get("Footprint", raw.get("Device", "-"))))
            footprint.setToolTip(footprint.text())
            for column, item in enumerate((check, designation, coordinate, footprint)):
                self.table.setItem(table_row, column, item)
            self.table.setRowHeight(table_row, 43)
        self._updating_table = False
        self.summary_label.setText(f"{sum(p.row_id in self.selected_ids for p in visible_points)} / {len(visible_points)} 可见")

    def table_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_table or item.column() != 0:
            return
        row_id = item.data(Qt.ItemDataRole.UserRole)
        target = self.selected_pad_ids if self.mode_combo.currentIndex() == 1 else self.selected_ids
        if item.checkState() == Qt.CheckState.Checked:
            target.add(row_id)
        else:
            target.discard(row_id)
        self.refresh()

    def select_visible(self, selected: bool) -> None:
        if self.mode_combo.currentIndex() == 1:
            for pad in self.visible_pads():
                self.selected_pad_ids.add(pad.pad_id) if selected else self.selected_pad_ids.discard(pad.pad_id)
        else:
            for point in self.visible_points():
                self.selected_ids.add(point.row_id) if selected else self.selected_ids.discard(point.row_id)
        self.refresh()

    def generate_gcode(self, points: list[PointData] | None = None) -> str:
        if self.mode_combo.currentIndex() == 1:
            return self.generate_pad_gcode()
        return self.generate_point_gcode(points)

    def generate_point_gcode(self, points: list[PointData] | None = None) -> str:
        points = self.ordered_selected_points() if points is None else points
        lines = ["; Banux PCB solder paste program", "; Source: PCB coordinate workbook",
                 f"; Points: {len(points)}", "M17", "G90", "G92 X0 Y0 Z0 E0",
                 f"G0 Z{gcode_number(self.safe_z.value())} F{gcode_number(self.z_feed.value())}"]
        for index, point in enumerate(points, 1):
            label = gcode_label(point.raw.get("Designator", "point"), 40)
            lines.extend([f"; {index} {label}",
                          f"G0 X{gcode_number(point.x)} Y{gcode_number(point.y)} F{gcode_number(self.travel_feed.value())}",
                          f"G1 Z{gcode_number(self.paste_z.value())} F{gcode_number(self.z_feed.value())}", "G91"])
            if self.extrude.value() > 0:
                lines.append(f"G1 E{gcode_number(self.extrude.value())} F{gcode_number(self.e_feed.value())}")
            if self.retract.value() > 0:
                lines.append(f"G1 E-{gcode_number(self.retract.value())} F{gcode_number(self.e_feed.value())}")
            lines.extend(["G90", f"G0 Z{gcode_number(self.safe_z.value())} F{gcode_number(self.z_feed.value())}"])
        lines.extend(["M84", "; End"])
        return "\n".join(lines) + "\n"

    def generate_pad_gcode(self) -> str:
        pads = self.ordered_selected_pads()
        segments = self.pad_dispense_segments(pads)
        source = gcode_label(self.current_gerber.name if self.current_gerber else "Gerber", 46)
        lines = ["; Banux PCB solder paste program", f"; Source: {source}",
                 f"; Pads: {len(pads)}", f"; Segments: {len(segments)}",
                 "M17", "G90", "G92 X0 Y0 Z0 E0",
                 f"G0 Z{gcode_number(self.safe_z.value())} F{gcode_number(self.z_feed.value())}"]
        for index, segment in enumerate(segments, 1):
            sx, sy = segment.start
            ex, ey = segment.end
            label = gcode_label(segment.label, 28)
            lines.extend([f"; {index} {label}",
                          f"G0 X{gcode_number(sx)} Y{gcode_number(sy)} F{gcode_number(self.travel_feed.value())}",
                          f"G1 Z{gcode_number(self.paste_z.value())} F{gcode_number(self.z_feed.value())}"])
            if segment.length <= 0.0005:
                lines.append("G91")
                if self.extrude.value() > 0:
                    lines.append(f"G1 E{gcode_number(self.extrude.value())} F{gcode_number(self.e_feed.value())}")
                if self.retract.value() > 0:
                    lines.append(f"G1 E-{gcode_number(self.retract.value())} F{gcode_number(self.e_feed.value())}")
                lines.append("G90")
            else:
                amount = max(0.0, segment.length * self.e_per_mm.value())
                lines.append("G92 E0")
                lines.append(f"G1 X{gcode_number(ex)} Y{gcode_number(ey)} E{gcode_number(amount, 4)} F{gcode_number(self.e_feed.value())}")
                if self.retract.value() > 0:
                    lines.extend(["G91",
                                  f"G1 E-{gcode_number(self.retract.value())} F{gcode_number(self.e_feed.value())}",
                                  "G90"])
            lines.append(f"G0 Z{gcode_number(self.safe_z.value())} F{gcode_number(self.z_feed.value())}")
        lines.extend(["M84", "; End"])
        return "\n".join(lines) + "\n"

    def path_stats(self, points: list[PointData],
                   segments: list[DispenseSegment] | None = None) -> tuple[float, float]:
        if segments is not None:
            travel = 0.0
            current: tuple[float, float] | None = None
            for segment in segments:
                if current is not None:
                    travel += math.hypot(segment.start[0] - current[0], segment.start[1] - current[1])
                current = segment.end
            dispense = sum(segment.length for segment in segments)
            z_distance = len(segments) * abs(self.safe_z.value() - self.paste_z.value()) * 2
            e_distance = dispense * self.e_per_mm.value() + len(segments) * self.retract.value()
            seconds = (travel / self.travel_feed.value() + z_distance / self.z_feed.value()
                       + e_distance / self.e_feed.value() + dispense / self.e_feed.value()) * 60
            return travel, seconds
        distance = sum(math.hypot(b.x - a.x, b.y - a.y) for a, b in zip(points, points[1:]))
        z_distance = len(points) * abs(self.safe_z.value() - self.paste_z.value()) * 2
        e_distance = len(points) * (self.extrude.value() + self.retract.value())
        seconds = (distance / self.travel_feed.value() + z_distance / self.z_feed.value()
                   + e_distance / self.e_feed.value()) * 60
        return distance, seconds

    def copy_gcode(self) -> None:
        QApplication.clipboard().setText(self.gcode_preview.toPlainText())
        self.statusBar().showMessage("G-code 已复制", 2500)

    def export_gcode(self) -> None:
        if self.mode_combo.currentIndex() == 1:
            if not self.selected_pad_ids:
                QMessageBox.warning(self, "无法导出", "请至少选择一个 Gerber 焊盘。")
                return
        elif not self.selected_ids:
            QMessageBox.warning(self, "无法导出", "请至少选择一个点胶坐标。")
            return
        source_file = self.current_gerber if self.mode_combo.currentIndex() == 1 else self.current_file
        default = (source_file.parent if source_file else ROOT) / "pcb_solder_paste.gcode"
        file_name, _ = QFileDialog.getSaveFileName(self, "导出 G-code", str(default),
                                                   "G-code (*.gcode *.nc);;文本文件 (*.txt)")
        if not file_name:
            return
        try:
            Path(file_name).write_text(self.generate_gcode(), encoding="utf-8", newline="\n")
        except OSError as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.statusBar().showMessage(f"已导出 {file_name}", 4000)

    def closeEvent(self, event) -> None:  # noqa: N802
        self.disconnect_wireless_device(show_status=False)
        super().closeEvent(event)

    def load_saved_device(self) -> WirelessDevice | None:
        """读取上一次记住的设备。兼容两种旧格式：
        1) 新版 6 字段：name/ip/bridge_port/http_port/wifi/ap
        2) 旧版字段名：device_name/device_ip/port/bridge/last_ip 等都接受
        """
        try:
            data = json.loads(DEVICE_CONFIG_PATH.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return None
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            return None

        def _first(*keys: str, default: Any = "") -> Any:
            for k in keys:
                if k in data and data[k] not in (None, ""):
                    return data[k]
            return default

        ip = str(_first("ip", "device_ip", "last_ip", "host", default=""))
        name = str(_first("name", "device_name", default=DEFAULT_DEVICE_NAME))
        if not ip:
            return None

        try:
            bp = int(_first("bridge_port", "bridge", "port", default=BRIDGE_PORT))
        except (TypeError, ValueError):
            bp = BRIDGE_PORT
        try:
            hp = int(_first("http_port", "http", default=80))
        except (TypeError, ValueError):
            hp = 80
        wifi = str(_first("wifi", "ssid", "last_ssid", default=""))
        ap_raw = _first("ap", "is_ap", "mode", default=False)
        if isinstance(ap_raw, bool):
            ap = ap_raw
        else:
            s = str(ap_raw).lower()
            ap = s in {"1", "true", "yes", "on", "ap"}
        dev = WirelessDevice(
            name=name or DEFAULT_DEVICE_NAME,
            ip=ip,
            bridge_port=bp or BRIDGE_PORT,
            http_port=hp or 80,
            wifi=wifi,
            ap=ap,
        )
        # 旧格式读完立刻转存新格式（下次就不用兼容判断）
        self.save_device(dev)
        return dev

    def save_device(self, device: WirelessDevice) -> None:
        self.saved_device = device
        payload = {
            "name": device.name,
            "ip": device.ip,
            "bridge_port": device.bridge_port,
            "http_port": device.http_port,
            "wifi": device.wifi,
            "ap": device.ap,
        }
        # 扩展：保存静态 IP / Topic（如果 dataclass 里有，或者通过 attribute 动态挂的）
        for extra in ("static_ip", "topic", "uid", "device_type"):
            val = getattr(device, extra, None)
            if val:
                payload[extra] = val
        try:
            DEVICE_CONFIG_PATH.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except OSError as exc:
            self.log_connection(f"save device failed: {exc}")

    def has_prepared_job(self) -> bool:
        if self.mode_combo.currentIndex() == 1:
            return bool(self.current_gerber and self.selected_pad_ids and self.gcode_preview.toPlainText().strip())
        return bool(self.current_file and self.selected_ids and self.gcode_preview.toPlainText().strip())

    def update_start_enabled(self) -> None:
        enabled = self.has_prepared_job()
        self.start_action.setEnabled(enabled)
        self.go_start_button.setEnabled(enabled)
        self.update_device_actions()

    def set_device_state(self, state: str, detail: str = "") -> None:
        if state == "online":
            self.device_status_label.setText("在线")
            self.device_status_label.setObjectName("deviceOnline")
        elif state == "setup":
            self.device_status_label.setText("待配网")
            self.device_status_label.setObjectName("deviceSetup")
        else:
            self.device_status_label.setText("离线")
            self.device_status_label.setObjectName("deviceOffline")
        self.device_status_label.style().unpolish(self.device_status_label)
        self.device_status_label.style().polish(self.device_status_label)
        self.device_info_label.setText(detail)
        self.update_device_actions()

    def update_device_actions(self) -> None:
        online = self.bridge_socket is not None
        prepared = self.has_prepared_job() if hasattr(self, "gcode_preview") else False
        busy = getattr(self, "_transfer_busy", False)
        if hasattr(self, "upload_button"):
            self.upload_button.setEnabled(online and prepared and not busy)
            self.execute_button.setEnabled(online and prepared and not busy)
        for button in getattr(self, "manual_buttons", []):
            button.setEnabled(online and not busy)

    def show_start_page(self) -> None:
        if not self.has_prepared_job():
            QMessageBox.information(self, "还不能开始", "请先导入 Gerber 并选择需要铺膏的焊盘。")
            self.update_start_enabled()
            return
        self.pages.setCurrentWidget(self.start_page)
        self.statusBar().showMessage("设备会自动连接，在线后即可传输或手动操作", 2500)
        self.poll_device_status()

    def update_remote_path_root(self, root: str) -> None:
        current = self.remote_path_edit.text().strip()
        filename = Path(current).name or "pcb_solder_paste.gcode"
        self.remote_path_edit.setText(f"{root}/{filename}")

    def log_connection(self, text: str) -> None:
        self.connection_log.appendPlainText(text.rstrip())
        self.connection_log.verticalScrollBar().setValue(
            self.connection_log.verticalScrollBar().maximum())

    @staticmethod
    def parse_discovery_packet(data: bytes, source_ip: str) -> WirelessDevice | None:
        try:
            text = data.decode("utf-8", errors="replace").strip()
        except UnicodeError:
            return None
        if not text.upper().startswith("BANPCBTOOL"):
            return None
        values: dict[str, str] = {}
        for token in text.split()[1:]:
            if "=" in token:
                key, value = token.split("=", 1)
                values[key.lower()] = value
        return WirelessDevice(
            name=values.get("name", "BanPCBTool"),
            ip=values.get("ip", source_ip),
            bridge_port=int(values.get("bridge", str(BRIDGE_PORT)) or BRIDGE_PORT),
            http_port=int(values.get("http", "80") or 80),
            wifi=values.get("wifi", ""),
            ap=values.get("ap", "0") in {"1", "true", "yes", "on"},
        )

    def refresh_device_combo(self) -> None:
        self.device_combo.clear()
        for device in self.devices:
            suffix = " AP" if device.ap else ""
            self.device_combo.addItem(
                f"{device.name}  {device.ip}:{device.bridge_port}{suffix}", device)

    def selected_wireless_device(self) -> WirelessDevice | None:
        if self.connected_device:
            return self.connected_device
        data = self.device_combo.currentData()
        if isinstance(data, WirelessDevice):
            return data
        if self.devices:
            return self.devices[0]
        return self.saved_device

    def discover_wireless_devices(self, include_ap_fallback: bool = False,
                                  log: bool = False,
                                  rounds: int = 3,
                                  extra_broadcast_targets: list[str] | None = None,
                                  recv_rounds_mult: int = 1) -> list[WirelessDevice]:
        """UDP 发现 BanPCBTool 设备。

        默认向每个 NIC 子网的定向广播 + 255.255.255.255 + 192.168.4.255 发 3 轮；
        recv_rounds_mult > 1 时会增加接收窗口，用于配网后密集等待上线。
        """
        if log:
            self.log_connection("search: UDP broadcast BANPCBTOOL?")
        found: dict[str, WirelessDevice] = {}
        udp = None
        try:
            targets = get_local_broadcast_addresses()
            if extra_broadcast_targets:
                targets = list(dict.fromkeys(targets + extra_broadcast_targets))
            udp = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            udp.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            udp.settimeout(0.25)
            for _round in range(max(1, int(rounds))):
                for bcast in targets:
                    try:
                        udp.sendto(b"BANPCBTOOL?", (bcast, DISCOVERY_PORT))
                    except OSError:
                        continue
            for _ in range(8 * max(1, int(recv_rounds_mult))):
                try:
                    data, address = udp.recvfrom(512)
                except (socket.timeout, TimeoutError):
                    QApplication.processEvents()
                    continue
                device = self.parse_discovery_packet(data, address[0])
                if device:
                    found[device.ip] = device
        except OSError as exc:
            if log:
                self.log_connection(f"search error: {exc}")
        finally:
            if udp is not None:
                try:
                    udp.close()
                except Exception:
                    pass
        if include_ap_fallback and DEFAULT_DEVICE_IP not in found:
            fallback = WirelessDevice("BanPCBTool", DEFAULT_DEVICE_IP, ap=True)
            found[DEFAULT_DEVICE_IP] = fallback
            if log:
                self.log_connection(
                    f"fallback: {fallback.name} {DEFAULT_DEVICE_IP} (AP 固定 IP)")
        self.devices = list(found.values())
        self.refresh_device_combo()
        if log:
            for device in self.devices:
                mode = "AP" if device.ap else "STA"
                self.log_connection(f"found: {device.name} {device.ip}:{device.bridge_port} {mode} wifi={device.wifi or '-'}")
        return self.devices

    def select_fixed_ip_device(self) -> None:
        for index in range(self.device_combo.count()):
            device = self.device_combo.itemData(index)
            if isinstance(device, WirelessDevice) and device.ip == DEFAULT_DEVICE_IP:
                self.device_combo.setCurrentIndex(index)
                break

    def pick_auto_device(self, log: bool = False) -> WirelessDevice | None:
        """发现优先级：UDP 广播发现(3 轮，AP 模式回落兜底) → saved_device。"""
        devices = self.discover_wireless_devices(
            include_ap_fallback=True, log=log, rounds=3, recv_rounds_mult=2)
        if devices:
            saved = self.saved_device
            best: WirelessDevice | None = None
            # 先 STA 模式，优先匹配 saved_device 名字
            for d in devices:
                if d.ap:
                    continue
                if saved and d.name == (saved.name or DEFAULT_DEVICE_NAME):
                    best = d
                    break
                if best is None:
                    best = d
            if best is not None:
                return best
            for d in devices:
                if d.ip != DEFAULT_DEVICE_IP:
                    return d
            return devices[0]
        return self.saved_device

    def query_bridge_info(self) -> dict[str, Any]:
        if not self.bridge_socket:
            return {}
        try:
            self.bridge_socket.sendall(b"@BPC STATUS\r\n")
            status_text = self.read_bridge_response(0.8)
        except OSError:
            return {}
        if "{" not in status_text:
            return {}
        try:
            return json.loads(status_text[status_text.index("{"):status_text.rindex("}") + 1])
        except (ValueError, json.JSONDecodeError):
            return {}

    def mark_connected(self, device: WirelessDevice, info: dict[str, Any] | None = None) -> None:
        info = info or {}
        device.name = str(info.get("name", device.name or DEFAULT_DEVICE_NAME))
        device.ip = str(info.get("ip", device.ip))
        device.bridge_port = int(info.get("bridge", device.bridge_port))
        device.http_port = int(info.get("http", device.http_port))
        device.wifi = str(info.get("ssid", device.wifi or ""))
        device.ap = bool(info.get("ap", device.ap))
        self.connected_device = device
        self.device_name_label.setText(device.name)
        self.save_device(device)
        if device.ap and not bool(info.get("wifi", not device.ap)):
            self.set_device_state("setup", f"{device.name} 已连接，等待配网")
        else:
            address = f"{device.ip}:{device.bridge_port}"
            self.set_device_state("online", f"{device.name} 已在线  {address}")

    def mark_offline(self, detail: str = "设备离线，开机后会自动重连") -> None:
        self.connected_device = None
        if self.bridge_socket:
            try:
                self.bridge_socket.close()
            except OSError:
                pass
        self.bridge_socket = None
        name = self.saved_device.name if self.saved_device else DEFAULT_DEVICE_NAME
        self.device_name_label.setText(name)
        self.set_device_state("offline", detail)

    def try_open_device(self, device: WirelessDevice, timeout: float = 0.9) -> bool:
        if not device or not device.ip:
            return False
        try:
            sock = self._tcp_connect(device.ip, device.bridge_port, timeout=timeout)
            self.bridge_socket = sock
            sock.sendall(b"@BPC PING\r\n")
            response = self.read_bridge_response(0.7)
            if "@BPC PONG" not in response and "@BPC OK" not in response:
                sock.sendall(b"@BPC HELLO\r\n")
                response += "\n" + self.read_bridge_response(0.7)
            if "@BPC PONG" not in response and "@BPC OK" not in response:
                raise OSError("设备握手失败")
            info = self.query_bridge_info()
            self.mark_connected(device, info)
            return True
        except OSError as exc:
            self.log_connection(f"auto connect failed: {exc}")
            self.mark_offline()
            return False

    def poll_device_status(self) -> None:
        """设备状态轮询（每 3 秒一次，首次启动 900ms 后执行）：
        1) 已有 TCP 桥接 → @BPC PING 心跳确认；掉线则 mark_offline
        2) 有 saved_device（用户上次成功连接过的 IP）→ 先直接 TCP saved_device:bridge_port
           （跳过 UDP 广播，保证即便路由器屏蔽 255.255.255.255 也能秒重连）
        3) saved_device 不通 → UDP 广播(3 轮 + 子网定向广播 + AP 兜底) + 自动挑选一台连接
        """
        if (getattr(self, "_connecting", False)
                or getattr(self, "_transfer_busy", False)
                or getattr(self, "_terminal_busy", False)):
            return

        # —— 1) 心跳检测已有连接 ——
        if self.bridge_socket:
            alive = False
            try:
                self.bridge_socket.sendall(b"@BPC PING\r\n")
                response = self.read_bridge_response(0.6)
                if "@BPC PONG" in response or "@BPC OK" in response:
                    alive = True
                    if self.connected_device:
                        info = self.query_bridge_info()
                        if info:
                            self.mark_connected(self.connected_device, info)
            except OSError:
                alive = False
            if not alive:
                self.mark_offline()
            else:
                return

        # —— 2) saved_device 直连 TCP 优先（不必等 UDP 广播）——
        sd = self.saved_device
        if sd and sd.ip and not sd.ap and sd.ip != DEFAULT_DEVICE_IP:
            try:
                if self.try_open_device(sd, timeout=1.6):
                    return
            except Exception as _ex:  # noqa: BLE001
                self.log_connection(f"saved_device TCP {sd.ip} failed: {_ex}")

        # —— 3) UDP 广播发现 + 自动选 + TCP 建链 ——
        device = self.pick_auto_device(log=False)
        if device:
            try:
                self.try_open_device(device, timeout=1.0)
            except Exception as _ex:  # noqa: BLE001
                self.log_connection(f"auto-pick connect failed: {_ex}")

    def try_auto_connect_hotspot(self) -> bool:
        """扫描附近 WiFi，自动把电脑 Wi-Fi 切到模块热点；成功返回 True。
        对齐 Ban-IOT APP：热点前缀 = DEFAULT_DEVICE_NAME，密码固定 = 12345678。
        连接热点前记住当前连接的 WiFi（存入 self._prev_ssid），配网完成后会自动切回。
        """
        self._prev_ssid = current_wifi_ssid()
        target = ""
        try:
            for ssid in scan_local_wifi():
                if ssid.lower().startswith(DEFAULT_DEVICE_NAME.lower()):
                    target = ssid
                    break
        except Exception:
            target = ""
        if not target:
            self.statusBar().showMessage("未扫描到模块热点", 3000)
            QMessageBox.information(
                self, "未找到模块热点",
                f"附近没扫到 {DEFAULT_DEVICE_NAME}XXXX 热点。\n\n"
                "请确认模块已通电、无线网卡已开启后重试；\n"
                "如果模块 STA 已经连到家庭 WiFi，请直接 UDP 发现后点\"首次设置\"。\n"
                "也可以手动在电脑 Wi-Fi 列表里连接该热点后再点\"首次设置\"\n"
                f"（热点名：{DEFAULT_DEVICE_NAME}XXXX，  密码：{DEFAULT_AP_PASSWORD}）。")
            return False
        self.statusBar().showMessage(f"正在自动连接热点 {target} (pwd {DEFAULT_AP_PASSWORD}) ...", 0)
        self.log_connection(f"auto connect hotspot: {target} pwd={DEFAULT_AP_PASSWORD}")
        # 优先用带密码的新协议；连接失败再尝试无密码（兼容旧固件未烧录的场景）
        if connect_to_hotspot(target, DEFAULT_AP_PASSWORD) or connect_to_hotspot(target, None):
            self.statusBar().showMessage(f"已连接模块热点 {target}", 2500)
            self.log_connection("hotspot connected")
            return True
        self.statusBar().showMessage("自动连接热点失败", 3000)
        QMessageBox.warning(
            self, "自动连接失败",
            f"尝试自动连接热点「{target}」失败。\n\n"
            f"请手动在电脑 Wi-Fi 列表中选择该热点（密码 {DEFAULT_AP_PASSWORD}）连接后，"
            "再点\"首次设置\"。")
        return False

    def start_first_setup(self) -> None:
        """Ban-IOT APP 风格的两步配网向导。
        Step 1：扫描 / 自动连接模块 AP（BanPCBToolXXXX, 12345678）
        Step 2：让模块扫描附近 WiFi，用户选择后 POST /config 保存，读取设备返回 JSON。
        """
        self._connecting = True
        try:
            # 如果已经在同一 WiFi 网内 UDP 发现了设备，直接弹出设置面板即可
            if self.bridge_socket and self.connected_device and not self.connected_device.ap:
                self.set_device_state("setup", f"{self.connected_device.name} 已在网内，弹出配网面板")
                self.show_provision_dialog(self.connected_device)
                return
            self.set_device_state("setup", "正在查找模块热点 (BanPCBToolXXXX)")
            # 1) 还没连上模块热点 → 尝试自动切过去
            curr = current_wifi_ssid()
            if not curr.lower().startswith(DEFAULT_DEVICE_NAME.lower()):
                if not self.try_auto_connect_hotspot():
                    return
            device = WirelessDevice(DEFAULT_DEVICE_NAME, DEFAULT_DEVICE_IP, ap=True)
            # 2) 尝试桥接握手；AP 模式下即便 TCP 桥接没通也能走 HTTP 配网，所以这里只是 try
            if self.try_open_device(device, timeout=2.0):
                self.show_provision_dialog(self.connected_device or device)
            else:
                # TCP 桥接不通但 HTTP 配网接口仍可用（AP 模式就是这个场景）
                self.log_connection("TCP bridge handshake skipped — use HTTP config via " + DEFAULT_DEVICE_IP)
                self.device_name_label.setText(device.name)
                self.set_device_state("setup", f"已连到 {DEFAULT_DEVICE_NAME} AP (HTTP 192.168.4.1)")
                self.show_provision_dialog(device)
        finally:
            self._connecting = False

    def show_provision_dialog(self, device: WirelessDevice | None = None) -> None:
        """Ban-IOT APP 风格配网面板（和 WifiConfigActivity 流程对齐）。
        - SSID 下拉框先用模块侧 /scan 结果（更准确，因为模块和要配的 WiFi 在同一物理空间），
          回退到本机扫描结果 + 手动输入。
        - POST /config 提交 ssid / pass / topic / static_ip / uid；
          返回 JSON 含 device_type / device_topic / static_ip / device_uid。
        """
        target = (device or self.selected_wireless_device()
                  or WirelessDevice(DEFAULT_DEVICE_NAME, DEFAULT_DEVICE_IP, ap=True))
        target_is_ap = bool(target.ap) or target.ip == DEFAULT_DEVICE_IP

        dlg = QDialog(self)
        dlg.setWindowTitle("WiFi 配网  ·  Ban-IOT 协议")
        dlg.setMinimumWidth(480)
        vbox = QVBoxLayout(dlg)

        banner = QLabel()
        banner.setStyleSheet(
            "padding:10px;border-radius:8px;font-weight:600;"
            + ("background:#fff3bf;color:#8a5a00;" if target_is_ap
               else "background:#d3f9d8;color:#087f5b;"))
        if target_is_ap:
            banner.setText(f"🔌 当前连接模块 AP · {target.name}  ·  192.168.4.1")
        else:
            banner.setText(f"🌐 已入网设备 · {target.name} @ {target.ip}")
        vbox.addWidget(banner)

        # —— Step 2 配置面板 ——
        title = QLabel("第 2 步 · 填写要让模块连接的 WiFi")
        title.setStyleSheet("font-weight:700;font-size:13px;margin-top:4px;")
        vbox.addWidget(title)

        form = QFormLayout()
        ssid_row = QHBoxLayout()
        scan_box = QComboBox()
        scan_box.setEditable(True)
        scan_box.setMinimumWidth(220)
        scan_btn = QPushButton("调模块扫描")
        local_btn = QPushButton("本机扫描")
        ssid_row.addWidget(scan_box, 1)
        ssid_row.addWidget(scan_btn)
        ssid_row.addWidget(local_btn)
        ssid_wrap = QWidget()
        ssid_wrap.setLayout(ssid_row)
        form.addRow("WiFi 名称 (SSID)", ssid_wrap)

        pass_edit = QLineEdit()
        pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        pass_tip = QLabel("空 = 开放 WiFi 或保留已保存的密码")
        pass_tip.setStyleSheet("color:#868e96;font-size:11px;")
        pass_wrap = QWidget()
        pv = QVBoxLayout(pass_wrap)
        pv.setContentsMargins(0, 0, 0, 0)
        pv.addWidget(pass_edit)
        pv.addWidget(pass_tip)
        form.addRow("WiFi 密码", pass_wrap)

        topic_edit = QLineEdit()
        topic_edit.setPlaceholderText("可选，例如 pcbtool001（用于云端/组群识别）")
        form.addRow("设备 Topic", topic_edit)

        static_ip_edit = QLineEdit()
        static_ip_edit.setPlaceholderText(
            "可选，留空自动推荐未占用静态IP（192.168.x.100~240）；填 DHCP 请手动删除")
        sip_row = QHBoxLayout()
        sip_auto_btn = QPushButton("🎲 推荐未占用IP")
        sip_check_btn = QPushButton("🔎 此IP冲突?")
        sip_row.addWidget(static_ip_edit, 1)
        sip_row.addWidget(sip_auto_btn)
        sip_row.addWidget(sip_check_btn)
        sip_wrap = QWidget()
        sip_wrap.setLayout(sip_row)
        form.addRow("静态 IP", sip_wrap)

        # —— 静态 IP 推荐 & 冲突检测助手 ——
        def _do_suggest_ip(_checked: bool = False) -> None:
            saved = [self.saved_device.ip] if (self.saved_device and self.saved_device.ip) else []
            sip_auto_btn.setEnabled(False)
            sip_auto_btn.setText("检测中…")
            QApplication.processEvents()

            def _worker() -> str:
                return suggest_static_ip(avoid_ips=saved, check_conflict=True)

            def _on_done(ip_: str) -> None:
                try:
                    static_ip_edit.setText(ip_)
                    prefixes = get_local_lan_prefixes()
                    if prefixes:
                        p, my = prefixes[0]
                        log(f"[√] 已推荐静态 IP：{ip_}    (当前PC {my}，网段 {p}.x)")
                    else:
                        log(f"[√] 已推荐静态 IP：{ip_}")
                finally:
                    sip_auto_btn.setEnabled(True)
                    sip_auto_btn.setText("🎲 推荐未占用IP")

            try:
                from concurrent.futures import ThreadPoolExecutor
                exe = getattr(self, "_one_off_pool", None)
                if exe is None:
                    exe = ThreadPoolExecutor(max_workers=1, thread_name_prefix="sip")
                    setattr(self, "_one_off_pool", exe)
                fut = exe.submit(_worker)

                def _poll() -> None:
                    if fut.done():
                        try:
                            _on_done(fut.result())
                        except Exception as _ex:  # noqa: BLE001
                            log(f"[×] 推荐静态IP失败：{_ex}")
                            sip_auto_btn.setEnabled(True)
                            sip_auto_btn.setText("🎲 推荐未占用IP")
                        return
                    QTimer.singleShot(150, _poll)
                _poll()
            except Exception as _ex:  # noqa: BLE001
                # 无并发支持就同步跑，几秒而已
                try:
                    _on_done(_worker())
                except Exception as _ex2:  # noqa: BLE001
                    log(f"[×] 推荐静态IP失败：{_ex} / {_ex2}")
                    sip_auto_btn.setEnabled(True)
                    sip_auto_btn.setText("🎲 推荐未占用IP")

        def _check_ip(_checked: bool = False) -> None:
            ip = static_ip_edit.text().strip()
            if not ip or "." not in ip:
                QMessageBox.information(dlg, "请先填 IP", "在「静态 IP」里填入地址，然后再测冲突。")
                return
            sip_check_btn.setEnabled(False)
            sip_check_btn.setText("检测中…")
            QApplication.processEvents()

            def _worker() -> bool:
                return is_ip_busy(ip)

            def _on_done(busy: bool) -> None:
                try:
                    if busy:
                        log(f"[×] 冲突：{ip} 已经被其他设备占用！建议点「🎲 推荐未占用IP」。")
                        QMessageBox.warning(dlg, "IP 已被占用",
                                            f"{ip} 已经有人在用，换一个或点推荐按钮。")
                    else:
                        log(f"[√] OK：{ip} 当前未占用，可以使用。")
                        QMessageBox.information(dlg, "IP 空闲", f"{ip} 当前无其他设备响应，可以使用。")
                finally:
                    sip_check_btn.setEnabled(True)
                    sip_check_btn.setText("🔎 此IP冲突?")

            try:
                from concurrent.futures import ThreadPoolExecutor
                exe = getattr(self, "_one_off_pool", None)
                if exe is None:
                    exe = ThreadPoolExecutor(max_workers=1, thread_name_prefix="sip")
                    setattr(self, "_one_off_pool", exe)
                fut = exe.submit(_worker)

                def _poll() -> None:
                    if fut.done():
                        try:
                            _on_done(fut.result())
                        except Exception as _ex:  # noqa: BLE001
                            log(f"[×] 冲突检测异常：{_ex}")
                            sip_check_btn.setEnabled(True)
                            sip_check_btn.setText("🔎 此IP冲突?")
                        return
                    QTimer.singleShot(150, _poll)
                _poll()
            except Exception as _ex:  # noqa: BLE001
                try:
                    _on_done(_worker())
                except Exception as _ex2:  # noqa: BLE001
                    log(f"[×] 冲突检测失败：{_ex} / {_ex2}")
                    sip_check_btn.setEnabled(True)
                    sip_check_btn.setText("🔎 此IP冲突?")

        sip_auto_btn.clicked.connect(_do_suggest_ip)
        sip_check_btn.clicked.connect(_check_ip)

        def _auto_fill_sip_if_empty() -> None:
            """GET /status 之后如果模块/PC都没填静态IP，自动帮用户推荐一个。"""
            if static_ip_edit.text().strip():
                return
            saved_list = [self.saved_device.ip] if (self.saved_device and self.saved_device.ip) else []
            try:
                ip_ = suggest_static_ip(avoid_ips=saved_list, check_conflict=True)
                static_ip_edit.setText(ip_)
                log(f"[→] 已自动推荐静态 IP {ip_}（可手动修改或点🎲再抽一次）")
            except Exception as _ex:  # noqa: BLE001
                log(f"[·] 自动推荐静态IP 跳过：{_ex}")

        uid_edit = QLineEdit()
        uid_edit.setPlaceholderText("可选，巴法云 / 其他云端的私钥 UID")
        uid_edit.setEchoMode(QLineEdit.EchoMode.Password)
        form.addRow("云端 UID", uid_edit)

        vbox.addLayout(form)

        # 操作按钮
        row_btn = QHBoxLayout()
        reconfig_btn = QPushButton("清除 WiFi 并重配")
        save_btn = QPushButton("保存并重启")
        save_btn.setStyleSheet(
            "padding:8px 14px;background:#087f5b;color:white;font-weight:700;border-radius:6px;")
        cancel_btn = QPushButton("关闭")
        row_btn.addWidget(reconfig_btn)
        row_btn.addStretch(1)
        row_btn.addWidget(cancel_btn)
        row_btn.addWidget(save_btn)
        vbox.addLayout(row_btn)

        result_box = QPlainTextEdit()
        result_box.setReadOnly(True)
        result_box.setPlaceholderText("保存结果会显示在这里……")
        result_box.setMaximumBlockCount(120)
        vbox.addWidget(result_box, 1)

        # --- helpers ---
        def log(msg: str) -> None:
            result_box.appendPlainText(msg)

        def temp_select_device() -> None:
            # 让 http_json 使用目标 device，而不是当前可能没建立 TCP 的 connected_device
            nonlocal _swap
            _swap = (self.connected_device, self.devices[:])
            self.connected_device = target

        def temp_restore_device() -> None:
            nonlocal _swap
            if _swap is None:
                return
            self.connected_device, self.devices = _swap
            _swap = None

        _swap: Any = None

        def fill_from_esp_scan() -> None:
            """GET /scan  -> 让模块自己扫描它能看到的 WiFi（Ban-IOT APP 的 scan 逻辑）。"""
            temp_select_device()
            current = scan_box.currentText().strip()
            try:
                log("[>] GET /scan (模块侧扫描) ...")
                result = self.http_json("/scan")
                names: list[str] = []
                if isinstance(result, list):
                    # 按 RSSI 降序排
                    def rssi(it: dict[str, Any]) -> int:
                        try:
                            return int(it.get("rssi", -100))
                        except (TypeError, ValueError):
                            return -100
                    sorted_list = sorted(result, key=rssi, reverse=True)
                    for item in sorted_list:
                        name = str(item.get("ssid", "")).strip()
                        if name and name not in names:
                            names.append(name)
                scan_box.clear()
                for n in names:
                    scan_box.addItem(n)
                if current:
                    idx = scan_box.findText(current)
                    if idx >= 0:
                        scan_box.setCurrentIndex(idx)
                    else:
                        scan_box.setEditText(current)
                log(f"[√] 模块扫描到 {len(names)} 个 WiFi，已填入下拉框")
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                log(f"[×] 模块侧扫描失败：{exc}")
                log("    提示：确认电脑已连到模块热点（或已在同一局域网）")
            finally:
                temp_restore_device()

        def fill_from_local_scan() -> None:
            current = scan_box.currentText().strip()
            scan_box.clear()
            names = scan_local_wifi()
            for n in names:
                scan_box.addItem(n)
            if current:
                idx = scan_box.findText(current)
                if idx >= 0:
                    scan_box.setCurrentIndex(idx)
                else:
                    scan_box.setEditText(current)
            log(f"本机扫描到 {len(names)} 个 WiFi（本机连 AP 模式时可能只有周边少量结果）")

        def save_and_reboot() -> None:
            """对齐 Ban-IOT APP sendConfigToDevice：POST /config + 读 JSON + 等待设备重启上线。
            保存成功后不会立刻关闭对话框，而是进入“等待设备上线”阶段：
              - PC 切回原家庭 WiFi（如果之前是连 AP 模式的话）
              - 密集做 UDP 发现（最多 35 秒），一旦找到 STA 模式新 IP：
                * 自动建立 TCP 桥接，顶部状态变绿 → 显示“在线 <IP>:<PORT>”
                * 把新 IP / name / wifi 存到 saved_device（下次启动直接连）
                * 对话框日志区打出 Web UI / OTA / 静态 IP 等所有操作入口
                * 弹「配网完成」面板给用户直接打开 Web 配置 / OTA / 完成

            全程 try/except BaseException，任何异常都写日志 + 弹窗，绝不冒泡到 Qt 事件循环闪退。
            """
            ssid = scan_box.currentText().strip()
            pwd = pass_edit.text()
            if not ssid:
                QMessageBox.information(dlg, "缺少 WiFi",
                                        "请先点击\"调模块扫描\"并选择要让模块连接的 WiFi。")
                return

            # —— UI 存活检测辅助：C++ 对象被删时，Qt/PyQt6 会抛 RuntimeError
            def _alive() -> bool:
                try:
                    # 对任意一个 widget 调属性，如果底层 C++ 对象已删就抛
                    _ = dlg.windowTitle()
                    _ = banner.text()
                    _ = save_btn.text()
                    return True
                except Exception:  # noqa: BLE001
                    return False

            dev_topic = ""
            dev_type = ""
            dev_sip = ""
            dev_uid = ""
            user_hint_static_ip = ""
            prev_ssid = ""

            try:
                # 1) 保存：优先 POST /config（Ban-IOT 协议），老固件 404 就 fallback /api/wifi
                temp_select_device()
                saved_ok = False
                try:
                    payload: dict[str, str] = {"ssid": ssid, "pass": pwd}
                    if topic_edit.text().strip():
                        payload["topic"] = topic_edit.text().strip()
                    if static_ip_edit.text().strip():
                        payload["static_ip"] = static_ip_edit.text().strip()
                    if uid_edit.text().strip():
                        payload["uid"] = uid_edit.text().strip()
                    log(f"[>] POST /config  {urllib.parse.urlencode(payload)}")
                    try:
                        resp = self.http_json("/config", payload)
                    except (urllib.error.HTTPError, OSError) as _ex1:
                        # 老固件（没有 /config 端点，一般是 HTTP 404 或空响应解析失败）
                        log(f"[·] 新接口 /config 不可用（{_ex1}），尝试旧版 /api/wifi ...")
                        legacy: dict[str, str] = {"ssid": ssid, "pass": pwd}
                        # 老固件不识别 static_ip/topic/uid，但仍然记一下让用户知道
                        resp = self.http_json("/api/wifi", legacy)
                    if isinstance(resp, dict):
                        log("[√] 模块保存成功，返回 JSON：")
                        log(json.dumps(resp, ensure_ascii=False, indent=2))
                        dev_topic = str(resp.get("device_topic", payload.get("topic", "")))
                        dev_type = str(resp.get("device_type", ""))
                        dev_uid = str(resp.get("device_uid", payload.get("uid", "")))
                        dev_sip = str(resp.get("static_ip", payload.get("static_ip", "")))
                    else:
                        log(f"[√] 保存成功（文本返回）：{resp}")
                    saved_ok = True
                except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                    log(f"[×] 保存失败：{exc}")
                    QMessageBox.warning(dlg, "保存失败", str(exc))
                finally:
                    temp_restore_device()
                if not saved_ok:
                    return

                # 2) 关闭输入，UI 进入“等待上线”阶段
                if not _alive():
                    return
                for w in (scan_box, scan_btn, local_btn, pass_edit,
                          topic_edit, static_ip_edit, uid_edit, save_btn,
                          reconfig_btn):
                    try:
                        w.setEnabled(False)
                    except Exception:
                        pass
                banner.setText(
                    f"📡 已保存 WiFi「{ssid}」，模块将在 3 秒后重启接入该网络…"
                    f" 等待设备在线中"
                )
                banner.setStyleSheet(
                    "padding:10px;border-radius:8px;font-weight:600;"
                    "background:#d0ebff;color:#1864ab;")
                save_btn.setText("已保存")
                cancel_btn.setText("关闭")
                try:
                    self.statusBar().showMessage("模块正在重启，等待上线…", 0)
                except Exception:
                    pass
                self.log_connection(
                    f"saved wifi {ssid}; waiting ESP reboot (3s) + STA connect")

                # 3) 先把 PC 切回原来的家庭 WiFi（等模块重启 3 秒 之后再切）
                prev_ssid = getattr(self, "_prev_ssid", "") or ""
                if prev_ssid.lower().startswith(DEFAULT_DEVICE_NAME.lower()):
                    # _prev_ssid 是在连 AP 之前记住的家庭 WiFi；不可能等于 AP 名称
                    # 如果这里是 AP，说明之前记录的值已经被污染；重置为空避免乱切
                    prev_ssid = ""
                user_hint_static_ip = static_ip_edit.text().strip() or dev_sip

            except BaseException as _boot_exc:  # noqa: BLE001
                tb = traceback.format_exc()
                log(f"[FATAL] save_and_reboot phase1 异常：{_boot_exc}\n{tb}")
                self.log_connection(f"[provision crash] phase1: {_boot_exc}")
                try:
                    QMessageBox.warning(dlg, "配网保存阶段出错",
                                        f"{_boot_exc}\n\n详细信息已写入日志面板。")
                except Exception:
                    pass
                return

            # —————— 异步阶段：4s 后开始切回家庭 WiFi + 探测设备 ——————
            wait_ctx: dict[str, Any] = {"attempt": 0, "max_attempts": 35, "stopped": False}

            def _after_reboot_phase() -> None:
                if wait_ctx["stopped"] or not _alive():
                    return
                try:
                    # 3a) 回连原家庭 WiFi（如果原来的是另一个非AP热点）
                    if prev_ssid and prev_ssid != ssid and current_wifi_ssid() != prev_ssid:
                        log(f"[WiFi] 把 PC 切回原 WiFi「{prev_ssid}」，准备寻找模块新地址…")
                        try:
                            reconnect_wifi(prev_ssid)
                        except Exception as _ex:  # noqa: BLE001
                            log(f"[WiFi] 回连失败：{_ex}（继续探测，可能 PC 已自动切换）")

                    # 3b) 启动密集等待上线的循环（异步，保持界面不卡）
                    wait_ctx["attempt"] = 0
                    wait_ctx["stopped"] = False
                    QTimer.singleShot(200, _probe_once)

                except BaseException as _exc:  # noqa: BLE001
                    tb = traceback.format_exc()
                    log(f"[FATAL] _after_reboot_phase 异常：{_exc}\n{tb}")
                    self.log_connection(f"[provision crash] reboot phase: {_exc}")
                    try:
                        QMessageBox.warning(dlg, "等待上线阶段出错",
                                            f"{_exc}\n\n详细信息已写入日志面板。")
                    except Exception:
                        pass

            def _probe_once() -> None:
                if wait_ctx["stopped"] or not _alive():
                    wait_ctx["stopped"] = True
                    return
                try:
                    wait_ctx["attempt"] = int(wait_ctx["attempt"]) + 1
                    attempt = int(wait_ctx["attempt"])
                    max_attempts = int(wait_ctx["max_attempts"])

                    # 组装额外的广播目标：如果用户指定了 static IP，把该子网广播加入
                    extra_targets: list[str] | None = None
                    if (user_hint_static_ip and "." in user_hint_static_ip
                            and user_hint_static_ip != "0.0.0.0"):
                        parts = user_hint_static_ip.rsplit(".", 1)
                        if len(parts) == 2 and parts[0]:
                            extra_targets = [f"{parts[0]}.255"]

                    sta_devices = self.discover_wireless_devices(
                        include_ap_fallback=False, log=False,
                        rounds=2, recv_rounds_mult=2,
                        extra_broadcast_targets=extra_targets,
                    )

                    # 优先选 STA 模式（不是 AP），优先匹配 static IP 或已保存 name
                    online: WirelessDevice | None = None
                    for candidate in sta_devices:
                        if candidate.ap:
                            continue
                        if user_hint_static_ip and candidate.ip == user_hint_static_ip:
                            online = candidate
                            break
                        if online is None:
                            online = candidate
                            continue
                        # 与 saved_device.name 匹配的优先
                        if self.saved_device and candidate.name == (
                                self.saved_device.name or DEFAULT_DEVICE_NAME):
                            online = candidate
                    if online is None:
                        for c in sta_devices:
                            if c.ip != DEFAULT_DEVICE_IP:
                                online = c
                                break

                    banner.setText(
                        f"📡 等待模块上线（第 {attempt}/{max_attempts} 次）… "
                        f"已发现 {len(sta_devices)} 台设备"
                    )

                    if online is None and attempt < max_attempts:
                        QTimer.singleShot(1000, _probe_once)
                        return

                    if online is None:
                        # 多次探测都没找到
                        banner.setText("⚠️ 未通过 UDP 找到模块，请直接用路由器管理页查看设备 IP")
                        banner.setStyleSheet(
                            "padding:10px;border-radius:8px;font-weight:600;"
                            "background:#fff5f5;color:#c92a2a;")
                        msg = (
                            "UDP 广播未在局域网内发现新设备，但模块配网过程已经成功。\n\n"
                            f"请用以下任一方法继续：\n"
                            f"  ① 登录 {ssid} 路由器管理页 → 查看设备列表 → 找到 "
                            f"ESP8266 或 BanPCBTool  → 记录 IP\n"
                            f"  ② 或等模块 AP 192.168.4.1 再次出现（STA 失败自动回落）\n"
                            f"  ③ 如有静态 IP：{user_hint_static_ip or '未设置'}\n"
                        )
                        log("[×] 等待上线超时，用户需手动查找 IP")
                        if user_hint_static_ip and user_hint_static_ip != "0.0.0.0":
                            try:
                                tmp = WirelessDevice(
                                    name=(self.saved_device.name
                                          if self.saved_device else None)
                                         or DEFAULT_DEVICE_NAME,
                                    ip=user_hint_static_ip, wifi=ssid, ap=False,
                                )
                                self.save_device(tmp)
                                self.log_connection(
                                    f"saved_device via static_ip: {tmp.ip}")
                            except Exception as _ex:  # noqa: BLE001
                                log(f"[×] 写入 saved_device(static_ip) 失败：{_ex}")
                        wait_ctx["stopped"] = True
                        try:
                            QMessageBox.information(dlg, "配网完成", msg)
                        except Exception:
                            pass
                        try:
                            dlg.accept()
                        except Exception:
                            pass
                        return

                    # 4) 找到在线设备 → 尝试建立 TCP 桥接 + 标记在线
                    log(f"[√] 发现模块上线：{online.name} @ {online.ip}  (AP={online.ap})")
                    success = False
                    try:
                        if self.try_open_device(online, timeout=1.5):
                            success = True
                        elif not online.ap:
                            try:
                                self.save_device(online)
                            except Exception:
                                pass
                            success = self.try_open_device(online, timeout=2.2)
                    except Exception as _ex:  # noqa: BLE001
                        log(f"[×] try_open_device 失败：{_ex}")
                        success = False

                    # 5) 更新 saved_device，下次启动直接用新 IP
                    try:
                        if not online.wifi:
                            online.wifi = ssid
                        self.save_device(online)
                        self.log_connection(
                            f"saved_device updated: {online.name} {online.ip}")
                    except Exception as _ex:  # noqa: BLE001
                        log(f"[×] save_device 失败：{_ex}")

                    # 6) 汇总日志 + 弹窗显示所有入口
                    http_url = f"http://{online.ip}:{online.http_port}/"
                    update_url = f"http://{online.ip}:{online.http_port}/update"
                    log("")
                    log("══════════════  配网完成 设备上线  ══════════════")
                    log(f"  设备名称 : {online.name}")
                    log(f"  设备 IP  : {online.ip}")
                    log(f"  网桥入口 : TCP {online.ip}:{online.bridge_port}")
                    log(f"  Web 配置 : {http_url}")
                    log(f"  OTA 固件 : {update_url}")
                    log(f"  所在 WiFi: {online.wifi or ssid}")
                    if dev_topic:
                        log(f"  设备Topic: {dev_topic}")
                    if dev_type:
                        log(f"  设备类型 : {dev_type}")
                    if dev_sip and dev_sip not in ("0.0.0.0", "255.255.255.255"):
                        log(f"  静态 IP  : {dev_sip}")
                    if dev_uid:
                        log(f"  UID      : {dev_uid[:8]}…")
                    log("═══════════════════════════════════════════════")

                    banner.setText(
                        f"✅ {online.name} 已在线 ｜ {online.ip}:{online.bridge_port} ｜ WiFi: "
                        f"{online.wifi or ssid}"
                    )
                    banner.setStyleSheet(
                        "padding:10px;border-radius:8px;font-weight:600;"
                        "background:#d3f9d8;color:#087f5b;")
                    try:
                        self.statusBar().showMessage(
                            f"{online.name} 已在线 {online.ip}:{online.bridge_port}", 4000)
                    except Exception:
                        pass

                    # 给用户一个最终操作面板：可以直接打开 Web UI / OTA，或者完成
                    final_dlg = QDialog(dlg)
                    final_dlg.setWindowTitle("配网完成")
                    v2 = QVBoxLayout(final_dlg)
                    bridge_html = ("<span style='color:#087f5b'>已自动连接</span>✅"
                                   if success else
                                   "<span style='color:#c92a2a'>未连通</span>"
                                   "（不影响 IP 写入，下次打开工具会继续尝试重连）")
                    info_txt = QLabel(
                        f"<b>{online.name}</b> 已上线并建立桥接：<br><br>"
                        f"&nbsp;&nbsp;🏠 WiFi ：{online.wifi or ssid}<br>"
                        f"&nbsp;&nbsp;🌐  IP   ：<b><code>{online.ip}</code></b><br>"
                        f"&nbsp;&nbsp;🔌 网桥 ：TCP {online.ip}:{online.bridge_port}<br>"
                        f"&nbsp;&nbsp;🌍 WebUI：{http_url}<br>"
                        f"&nbsp;&nbsp;⬆️ OTA  ：{update_url}<br>"
                        + (f"&nbsp;&nbsp;📛 Topic：{dev_topic}<br>" if dev_topic else "")
                        + (f"&nbsp;&nbsp;📛 Static：{dev_sip}<br>"
                           if dev_sip and dev_sip not in ("0.0.0.0", "255.255.255.255")
                           else "")
                        + f"<br>TCP 桥接：{bridge_html}"
                    )
                    info_txt.setTextFormat(Qt.TextFormat.RichText)
                    info_txt.setWordWrap(True)
                    info_txt.setStyleSheet(
                        "padding:14px;background:#f8f9fa;border-radius:8px;")
                    v2.addWidget(info_txt)

                    actions = QHBoxLayout()
                    open_web_btn = QPushButton("打开 Web 配置")
                    open_ota_btn = QPushButton("打开 OTA 升级")
                    ok_btn = QPushButton("完成")
                    ok_btn.setStyleSheet(
                        "padding:6px 12px;background:#087f5b;color:white;"
                        "font-weight:700;border-radius:4px;")
                    actions.addWidget(open_web_btn)
                    actions.addWidget(open_ota_btn)
                    actions.addStretch(1)
                    actions.addWidget(ok_btn)
                    v2.addLayout(actions)

                    def _open(u: str) -> None:
                        try:
                            webbrowser.open(u)
                        except Exception as _ex:  # noqa: BLE001
                            log(f"[×] 打开浏览器失败：{_ex}")
                            try:
                                QMessageBox.warning(final_dlg, "打开失败", str(_ex))
                            except Exception:
                                pass

                    open_web_btn.clicked.connect(lambda _c=False, u=http_url: _open(u))
                    open_ota_btn.clicked.connect(lambda _c=False, u=update_url: _open(u))
                    ok_btn.clicked.connect(final_dlg.accept)

                    final_dlg.resize(560, 300)
                    wait_ctx["stopped"] = True
                    try:
                        if final_dlg.exec() == 1:  # QDialog.DialogCode.Accepted == 1
                            dlg.accept()
                    except Exception:
                        pass

                except BaseException as _exc:  # noqa: BLE001
                    wait_ctx["stopped"] = True
                    tb = traceback.format_exc()
                    log(f"[FATAL] _probe_once 异常：{_exc}\n{tb}")
                    self.log_connection(f"[provision crash] probe: {_exc}")
                    try:
                        QMessageBox.warning(
                            dlg, "等待上线阶段出错",
                            f"{_exc}\n\n（已阻止闪退，详细信息见日志面板）")
                    except Exception:
                        # 如果连 dlg 都挂了，就只写主窗口日志
                        try:
                            self.statusBar().showMessage(
                                f"探测流程异常：{_exc}", 6000)
                        except Exception:
                            pass

            def _safe_reboot_phase() -> None:
                try:
                    _after_reboot_phase()
                except BaseException as _exc:  # noqa: BLE001
                    wait_ctx["stopped"] = True
                    tb = traceback.format_exc()
                    log(f"[FATAL] _after_reboot_phase 异常：{_exc}\n{tb}")
                    self.log_connection(f"[provision crash] reboot: {_exc}")
                    try:
                        QMessageBox.warning(dlg, "等待上线阶段出错",
                                            f"{_exc}\n\n（详细信息见日志面板，已阻止闪退）")
                    except Exception:
                        pass

            # 用户若在等待期间点关闭按钮：停止探测定时器
            def _on_close_clicked() -> None:
                wait_ctx["stopped"] = True
                try:
                    dlg.reject()
                except Exception:
                    pass

            try:
                cancel_btn.clicked.disconnect()
            except Exception:
                pass
            cancel_btn.clicked.connect(_on_close_clicked)

            # 模块 POST /config 返回后 3 秒后才重启，等待这个延迟后再开始扫
            QTimer.singleShot(4000, _safe_reboot_phase)

        def clear_and_restart() -> None:
            """POST /reconfig → 清除保存的 WiFi + 重启回到 AP 模式（Ban-IOT 协议）。"""
            ans = QMessageBox.question(
                dlg, "确认清除",
                "这会清除模块已保存的 WiFi，并强制重启回到 AP 模式（需要重新配网）。\n是否继续？")
            if ans != QMessageBox.StandardButton.Yes:
                return
            temp_select_device()
            try:
                log("[>] POST /reconfig (清除 WiFi + 重启 AP 模式) ...")
                result = self.http_json("/reconfig", {})
                log(f"[√] {result}")
                QMessageBox.information(dlg, "已清除",
                                        "模块已清除 WiFi，即将重启。\n"
                                        "重启后请重新执行向导，或手动从 WiFi 列表连模块热点："
                                        f" {DEFAULT_DEVICE_NAME}XXXX  (pwd {DEFAULT_AP_PASSWORD})")
                # 重新切到 AP，方便下一步再配
                if target_is_ap:
                    QTimer.singleShot(4000, lambda: self.try_auto_connect_hotspot())
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                log(f"[×] 清除失败：{exc}")
                QMessageBox.warning(dlg, "清除失败", str(exc))
            finally:
                temp_restore_device()

        def load_prefill_from_status() -> None:
            """GET /status → 预填已保存的 SSID / Topic / 静态 IP / UID（Ban-IOT APP 逻辑）。"""
            temp_select_device()
            try:
                log("[>] GET /status (读取模块状态与已保存字段)...")
                s = self.http_json("/status")
                if isinstance(s, dict):
                    log(json.dumps(s, ensure_ascii=False, indent=2))
                    saved_ssid = str(s.get("saved_ssid", ""))
                    saved_pass = str(s.get("saved_pass", ""))
                    saved_topic = str(s.get("device_topic", "")) or str(s.get("saved_topic", ""))
                    saved_sip = str(s.get("static_ip", ""))
                    saved_uid = str(s.get("device_uid", ""))
                    if saved_ssid and scan_box.findText(saved_ssid) < 0:
                        scan_box.insertItem(0, saved_ssid)
                        scan_box.setCurrentIndex(0)
                    if saved_pass:
                        pass_edit.setText(saved_pass)
                    if saved_topic:
                        topic_edit.setText(saved_topic)
                    if saved_sip and saved_sip not in ("", "0.0.0.0", "255.255.255.255"):
                        static_ip_edit.setText(saved_sip)
                    if saved_uid:
                        uid_edit.setText(saved_uid)
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                log(f"[×] GET /status 失败：{exc}")
                # GET /status 失败 → 可能是老固件（没新接口）。无论如何，用户要配网，
                # 仍然帮用户推荐一个静态 IP（老固件会存 static_ip=0=DHCP，但值已经POST过）
                QTimer.singleShot(100, _auto_fill_sip_if_empty)
            finally:
                temp_restore_device()

            # GET /status 成功但没 saved static_ip → 自动推荐
            if not static_ip_edit.text().strip():
                QTimer.singleShot(100, _auto_fill_sip_if_empty)

        scan_btn.clicked.connect(fill_from_esp_scan)
        local_btn.clicked.connect(fill_from_local_scan)
        save_btn.clicked.connect(save_and_reboot)
        reconfig_btn.clicked.connect(clear_and_restart)
        cancel_btn.clicked.connect(dlg.reject)

        # 打开后先试 GET /status 预填已有字段，再试一次 ESP 扫描
        QTimer.singleShot(0, load_prefill_from_status)
        if target_is_ap:
            QTimer.singleShot(300, fill_from_esp_scan)

        dlg.resize(520, 560)
        dlg.exec()

    def _tcp_connect(self, host: str, port: int, timeout: float = 2.5) -> socket.socket:
        """非阻塞 TCP 连接，连接等待期间保持界面响应；失败抛 OSError。"""
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.setblocking(False)
        deadline = time.monotonic() + timeout
        try:
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    raise TimeoutError(f"连接 {host}:{port} 超时")
                try:
                    sock.connect((host, port))
                except (BlockingIOError, InterruptedError):
                    pass  # 连接进行中，等待可写
                except OSError:
                    sock.close()
                    raise
                else:
                    sock.setblocking(True)
                    sock.settimeout(0.6)
                    return sock
                _, writable, _ = select.select([], [sock], [], min(0.1, remaining))
                if writable:
                    err = sock.getsockopt(socket.SOL_SOCKET, socket.SO_ERROR)
                    if err == 0:
                        sock.setblocking(True)
                        sock.settimeout(0.6)
                        return sock
                    sock.close()
                    raise OSError(err, os.strerror(err) or f"连接 {host}:{port} 失败")
                QApplication.processEvents()
        except BaseException:
            try:
                sock.close()
            except OSError:
                pass
            raise

    def connect_wireless_device(self) -> None:
        self.poll_device_status()

    def disconnect_wireless_device(self, show_status: bool = True) -> None:
        if self.bridge_socket:
            try:
                self.bridge_socket.close()
            except OSError:
                pass
        self.bridge_socket = None
        self.connected_device = None
        if show_status:
            self.device_info_label.setText("已断开")
            self.statusBar().showMessage("无线连接已断开", 2000)

    def read_bridge_response(self, timeout_s: float = 2.0) -> str:
        sock = self.bridge_socket
        if sock is None:
            return ""
        chunks: list[bytes] = []
        try:
            sock.settimeout(0.2)
        except OSError:
            return ""
        stop_at = time.monotonic() + timeout_s
        while time.monotonic() < stop_at:
            try:
                chunk = sock.recv(1024)
            except (socket.timeout, TimeoutError):
                QApplication.processEvents()
                continue
            except OSError:
                break
            if not chunk:
                break
            chunks.append(chunk)
            text = b"".join(chunks).decode("utf-8", errors="replace")
            if ("banux$ " in text or "OK block" in text or "OK clear" in text or
                    "@BPC PONG" in text or "@BPC OK" in text or "Error:" in text):
                return text.strip()
        return b"".join(chunks).decode("utf-8", errors="replace").strip()

    def send_bridge_command(self, command: str, timeout_s: float = 2.5) -> str:
        sock = self.bridge_socket
        if not sock:
            raise OSError("无线设备未连接")
        self.log_connection(f"> {command}")
        sock.sendall(command.encode("ascii") + b"\r\n")
        response = self.read_bridge_response(timeout_s)
        if response:
            self.log_connection(response)
        if "Error:" in response or "failed" in response.lower():
            raise OSError(response or "命令执行失败")
        return response

    def _transfer_ready(self) -> bool:
        if self.mode_combo.currentIndex() == 1 and not self.selected_pad_ids:
            QMessageBox.warning(self, "无法传输", "请至少选择一个 Gerber 焊盘。")
            return False
        if self.mode_combo.currentIndex() != 1 and not self.selected_ids:
            QMessageBox.warning(self, "无法传输", "请至少选择一个点胶坐标。")
            return False
        if not self.bridge_socket:
            QMessageBox.information(self, "未连接", "请先搜索并连接 BanPCBTool。")
            return False
        return True

    def upload_gcode_to_device(self) -> None:
        if not self._transfer_ready():
            return
        path = self.remote_path_edit.text().strip()
        if not path.startswith(("/flash/", "/sd/")) or " " in path:
            QMessageBox.warning(self, "路径无效", "路径必须是 /flash/... 或 /sd/...，且不能包含空格。")
            return
        payload = self.generate_gcode().encode("ascii")
        self._transfer_busy = True
        self.update_device_actions()
        try:
            self.send_bridge_command(f"recv -c {path}", 3.0)
            for offset in range(0, len(payload), UPLOAD_CHUNK_SIZE):
                chunk = payload[offset:offset + UPLOAD_CHUNK_SIZE]
                encoded = base64.b64encode(chunk).decode("ascii")
                self.send_bridge_command(f"recv -b {path} {offset} {encoded}", 3.0)
                percent = int(((offset + len(chunk)) * 100) / max(1, len(payload)))
                self.transfer_progress.setText(f"{percent}%  {offset + len(chunk)} / {len(payload)} bytes")
                QApplication.processEvents()
        except OSError as exc:
            self.mark_offline("设备连接中断")
            QMessageBox.warning(self, "传输失败", str(exc))
            return
        finally:
            self._transfer_busy = False
            self.update_device_actions()
        self.statusBar().showMessage(f"已传输到 {path}", 3500)

    def execute_remote_gcode(self) -> None:
        if not self.bridge_socket:
            QMessageBox.information(self, "未连接", "请先连接 BanPCBTool。")
            return
        path = self.remote_path_edit.text().strip()
        self._transfer_busy = True
        self.update_device_actions()
        try:
            self.send_bridge_command(f"gcode -f {path}", 4.0)
        except OSError as exc:
            self.mark_offline("设备连接中断")
            QMessageBox.warning(self, "执行失败", str(exc))
            return
        finally:
            self._transfer_busy = False
            self.update_device_actions()
        self.statusBar().showMessage(f"已发送执行命令：{path}", 3000)

    def send_manual_command(self, command: str, timeout_s: float = 2.5) -> bool:
        if not self.bridge_socket:
            QMessageBox.information(self, "未连接", "请先连接 BanPCBTool。")
            return False
        try:
            self.send_bridge_command(command, timeout_s)
        except OSError as exc:
            self.mark_offline("设备连接中断")
            QMessageBox.warning(self, "操作失败", str(exc))
            return False
        return True

    def show_bridge_terminal(self) -> None:
        """桥接命令行：把输入原样透传给 WiFi 模块，并实时显示模块回传的全部数据。

        模块会把非控制行原样转发到自己的串口（给 STM32），也会把 STM32 发给它的
        数据原样转发到 TCP，所以下位机回传的内容会实时出现在这个窗口里。

        打开期间必须暂停 device_timer 的心跳轮询（PING / STATUS），
        否则两边会争抢同一个 socket —— 轮询会把下位机的数据吞进自己的缓冲区。
        """
        if not self.bridge_socket:
            QMessageBox.information(self, "未连接", "请先连接 BanPCBTool 再打开命令行。")
            return
        if getattr(self, "_terminal_busy", False):
            return

        sock = self.bridge_socket
        device = self.connected_device
        if device:
            host_text = f"{device.name}  {device.ip}:{device.bridge_port}"
        else:
            host_text = "已连接设备"

        dialog = QDialog(self)
        dialog.setWindowTitle(f"命令行 · {host_text}")
        dialog.resize(780, 520)
        root = QVBoxLayout(dialog)

        tip = QLabel("输入内容原样发送给 WiFi 模块；模块侧非 @BPC 开头的行会转发给 STM32，"
                     "STM32 发给模块的数据也会实时显示在这里。")
        tip.setObjectName("muted")
        tip.setWordWrap(True)
        root.addWidget(tip)

        output = QPlainTextEdit()
        output.setReadOnly(True)
        output.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        output.setFont(QFont("Consolas", 10))
        root.addWidget(output, 1)

        input_row = QHBoxLayout()
        cmd_edit = QLineEdit()
        cmd_edit.setPlaceholderText("输入要发送的内容，回车发送（例如 G28 / M114）")
        send_btn = QPushButton("发送")
        input_row.addWidget(cmd_edit, 1)
        input_row.addWidget(send_btn)
        root.addLayout(input_row)

        opt_row = QHBoxLayout()
        crlf_box = QCheckBox("结尾加 \\r\\n")
        crlf_box.setChecked(True)
        hex_box = QCheckBox("HEX 显示")
        clear_btn = QPushButton("清空")
        close_btn = QPushButton("关闭")
        opt_row.addWidget(crlf_box)
        opt_row.addWidget(hex_box)
        opt_row.addStretch(1)
        opt_row.addWidget(clear_btn)
        opt_row.addWidget(close_btn)
        root.addLayout(opt_row)

        state = {"alive": True}

        def append_text(text: str) -> None:
            output.moveCursor(QTextCursor.MoveOperation.End)
            output.insertPlainText(text)
            output.verticalScrollBar().setValue(output.verticalScrollBar().maximum())

        def append_bytes(raw: bytes) -> None:
            if hex_box.isChecked():
                append_text(" ".join(f"{b:02X}" for b in raw) + " ")
            else:
                append_text(raw.decode("utf-8", errors="replace")
                               .replace("\r\n", "\n").replace("\r", "\n"))

        def note(text: str) -> None:
            append_text(f"\n{text}\n")

        timer = QTimer(dialog)

        def stop(reason: str = "") -> None:
            state["alive"] = False
            timer.stop()
            if reason:
                note(reason)

        def pump() -> None:
            """把 socket 上待收的数据全部搬到终端（非阻塞）。"""
            if not state["alive"]:
                return
            try:
                for _ in range(64):  # 每轮最多搬 64 包，避免大流量时卡住界面
                    readable, _, _ = select.select([sock], [], [], 0)
                    if not readable:
                        break
                    chunk = sock.recv(4096)
                    if not chunk:
                        stop("[连接已断开]")
                        self.mark_offline("设备连接中断")
                        return
                    append_bytes(chunk)
            except (BlockingIOError, InterruptedError):
                pass
            except (OSError, ValueError) as exc:
                stop(f"[连接已断开: {exc}]")
                self.mark_offline("设备连接中断")

        def do_send() -> None:
            text = cmd_edit.text()
            if not text:
                return
            payload = text.encode("utf-8") + (b"\r\n" if crlf_box.isChecked() else b"")
            try:
                sock.sendall(payload)
            except OSError as exc:
                stop(f"[发送失败: {exc}]")
                self.mark_offline("设备连接中断")
                return
            note(f"> {text}")
            cmd_edit.clear()

        timer.setInterval(80)
        timer.timeout.connect(pump)
        send_btn.clicked.connect(do_send)
        cmd_edit.returnPressed.connect(do_send)
        clear_btn.clicked.connect(output.clear)
        close_btn.clicked.connect(dialog.accept)

        self._terminal_busy = True
        try:
            note("[已进入透传模式，等待数据…]")
            timer.start()
            dialog.exec()
        finally:
            timer.stop()
            state["alive"] = False
            self._terminal_busy = False
            # 关闭后立刻刷新一次设备状态，不必等下一个 3 秒周期
            QTimer.singleShot(300, self.poll_device_status)

    def jog_axis(self, axis: str, direction: float) -> None:
        amount = self.extruder_step.value() if axis == "E" else self.jog_step.value()
        value = gcode_number(amount * direction, 4 if axis == "E" else 3)
        feed = gcode_number(self.jog_feed.value(), 0)
        if not self.send_manual_command("gcode G91", 1.5):
            return
        if not self.send_manual_command(f"gcode G1 {axis}{value} F{feed}", 3.0):
            return
        self.send_manual_command("gcode G90", 1.5)

    def http_json(self, path: str, data: dict[str, str] | None = None) -> Any:
        device = self.connected_device or self.selected_wireless_device()
        if not device:
            raise OSError("请先选择或连接设备")
        url = f"http://{device.ip}:{device.http_port}{path}"
        body = None
        headers = {}
        if data is not None:
            body = urllib.parse.urlencode(data).encode("utf-8")
            headers["Content-Type"] = "application/x-www-form-urlencoded"
        request = urllib.request.Request(url, data=body, headers=headers, method="POST" if data is not None else "GET")
        with urllib.request.urlopen(request, timeout=5.0) as response:
            text = response.read().decode("utf-8", errors="replace")
        return json.loads(text) if text.strip().startswith(("{", "[")) else text

    def show_wireless_settings(self) -> None:
        """无线设置（Ban-IOT 协议 + 老接口双兼容）。
        新增字段：Topic / 静态 IP / UID；
        新接口：GET /status, GET /scan, POST /config, POST /reconfig, POST /setip；
        保留老接口：/api/wifi, /api/ap, /api/clear, /api/restart, /api/ota 以兼容未烧新固件的模块。
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("无线设置  ·  Ban-IOT 协议")
        dialog.resize(700, 500)
        root_layout = QVBoxLayout(dialog)

        base_device = self.connected_device or self.selected_wireless_device()

        # ===== 主机 & 连接状态 =====
        conn = QGroupBox("连接")
        cf = QFormLayout(conn)
        host_edit = QLineEdit(base_device.ip if base_device else DEFAULT_DEVICE_IP)
        status_btn = QPushButton("读取状态")
        status_wrap = QWidget()
        sw = QHBoxLayout(status_wrap)
        sw.setContentsMargins(0, 0, 0, 0)
        sw.addWidget(host_edit, 1)
        sw.addWidget(status_btn)
        cf.addRow("模块 IP", status_wrap)
        root_layout.addWidget(conn)

        # ===== WiFi 配网（Ban-IOT /config 协议）=====
        prov = QGroupBox("WiFi 配网  (POST /config)")
        pf = QFormLayout(prov)

        # SSID 下拉框 + 本机扫描 + 模块扫描
        ssid_row = QHBoxLayout()
        scan_box = QComboBox()
        scan_box.setEditable(True)
        ssid_edit = QLineEdit()
        ssid_edit.setPlaceholderText("为空则用右侧扫描结果")
        esp_scan_btn = QPushButton("模块扫描")
        local_scan_btn = QPushButton("本机扫描")
        ssid_row.addWidget(scan_box, 1)
        ssid_row.addWidget(esp_scan_btn)
        ssid_row.addWidget(local_scan_btn)
        ssid_wrap = QWidget()
        ssid_wrap.setLayout(ssid_row)
        pf.addRow("SSID 扫描", ssid_wrap)
        pf.addRow("SSID 名称", ssid_edit)

        pass_edit = QLineEdit()
        pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        pf.addRow("WiFi 密码", pass_edit)

        topic_edit = QLineEdit()
        topic_edit.setPlaceholderText("可选，如 pcbtool001")
        pf.addRow("设备 Topic", topic_edit)

        sip_edit = QLineEdit()
        sip_edit.setPlaceholderText("可选，DHCP 留空；例如 192.168.1.200")
        pf.addRow("静态 IP", sip_edit)

        uid_edit = QLineEdit()
        uid_edit.setPlaceholderText("可选，云端 UID")
        uid_edit.setEchoMode(QLineEdit.EchoMode.Password)
        pf.addRow("云端 UID", uid_edit)

        prov_btns = QHBoxLayout()
        save_config_btn = QPushButton("保存配网 (/config)")
        save_config_btn.setStyleSheet(
            "padding:6px 10px;background:#087f5b;color:white;font-weight:700;border-radius:4px;")
        prov_btns.addWidget(save_config_btn)
        prov_btns.addStretch(1)
        pf.addRow(prov_btns)
        root_layout.addWidget(prov)

        # ===== 控制操作 =====
        ctl = QGroupBox("控制")
        cb = QHBoxLayout(ctl)
        ap_btn = QPushButton("开启AP模式")
        reconfig_btn = QPushButton("重置配网(/reconfig)")
        clear_btn = QPushButton("旧版清除WiFi")
        setip_btn = QPushButton("设置静态IP")
        restart_btn = QPushButton("重启模块")
        for b in (ap_btn, reconfig_btn, clear_btn, setip_btn, restart_btn):
            cb.addWidget(b)
        cb.addStretch(1)
        root_layout.addWidget(ctl)

        # ===== OTA =====
        ota = QGroupBox("OTA")
        of = QFormLayout(ota)
        ota_edit = QLineEdit()
        of.addRow("固件 URL", ota_edit)
        ob = QHBoxLayout()
        ota_btn = QPushButton("执行 OTA (/api/ota)")
        open_update_btn = QPushButton("浏览器打开 /update")
        ob.addWidget(ota_btn)
        ob.addWidget(open_update_btn)
        ob.addStretch(1)
        of.addRow(ob)
        root_layout.addWidget(ota)

        # ===== 日志窗口 =====
        info = QPlainTextEdit()
        info.setReadOnly(True)
        info.setFont(QFont("Consolas", 9))
        root_layout.addWidget(info, 1)
        close_btn = QPushButton("关闭")
        right = QHBoxLayout()
        right.addStretch(1)
        right.addWidget(close_btn)
        root_layout.addLayout(right)

        # ===== helpers =====
        def log(text: str) -> None:
            info.appendPlainText(text)

        def target_device() -> WirelessDevice:
            host = host_edit.text().strip()
            if not host:
                raise OSError("请先填写模块 IP")
            if base_device is not None:
                base_device.ip = host
                return base_device
            return WirelessDevice(DEFAULT_DEVICE_NAME, host)

        def call(path: str, payload: dict[str, str] | None = None) -> Any:
            old_cd = self.connected_device
            old_ds = self.devices[:]
            self.connected_device = target_device()
            try:
                result = self.http_json(path, payload)
            finally:
                self.connected_device = old_cd
                self.devices = old_ds
            if isinstance(result, (dict, list)):
                log(json.dumps(result, ensure_ascii=False, indent=2))
            else:
                log(str(result))
            return result

        def do_status() -> None:
            log("[>] GET /status")
            try:
                s = call("/status")
                if isinstance(s, dict):
                    # 预填表单
                    saved_ssid = str(s.get("saved_ssid", ""))
                    if saved_ssid and not ssid_edit.text().strip():
                        ssid_edit.setText(saved_ssid)
                        if scan_box.findText(saved_ssid) < 0:
                            scan_box.insertItem(0, saved_ssid)
                    saved_pass = str(s.get("saved_pass", ""))
                    if saved_pass and not pass_edit.text():
                        pass_edit.setText(saved_pass)
                    dev_topic = str(s.get("device_topic", "")) or str(s.get("saved_topic", ""))
                    if dev_topic and not topic_edit.text():
                        topic_edit.setText(dev_topic)
                    sip = str(s.get("static_ip", ""))
                    if sip and sip not in ("", "0.0.0.0", "255.255.255.255") and not sip_edit.text():
                        sip_edit.setText(sip)
                    uid = str(s.get("device_uid", ""))
                    if uid and not uid_edit.text():
                        uid_edit.setText(uid)
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                log(f"[×] /status 失败：{exc}")
                log("    尝试旧版 /api/info …")
                try:
                    call("/api/info")
                except Exception as exc2:  # noqa: BLE001
                    log(f"    [×] 也失败：{exc2}")

        def do_local_scan() -> None:
            names = scan_local_wifi()
            scan_box.clear()
            for n in names:
                scan_box.addItem(n)
            log(f"本机扫描到 {len(names)} 个 WiFi")

        def do_esp_scan() -> None:
            log("[>] GET /scan (模块侧)")
            try:
                result = call("/scan")
                names: list[str] = []
                if isinstance(result, list):
                    def rssi(it: dict[str, Any]) -> int:
                        try:
                            return int(it.get("rssi", -100))
                        except (TypeError, ValueError):
                            return -100
                    for item in sorted(result, key=rssi, reverse=True):
                        name = str(item.get("ssid", "")).strip()
                        if name and name not in names:
                            names.append(name)
                scan_box.clear()
                for n in names:
                    scan_box.addItem(n)
                log(f"模块扫描到 {len(names)} 个 WiFi")
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                log(f"[×] {exc}；回退 /api/scan")
                try:
                    result2 = call("/api/scan")
                    scan_box.clear()
                    if isinstance(result2, list):
                        for item in result2:
                            n = str(item.get("ssid", "")).strip()
                            if n:
                                scan_box.addItem(n)
                except Exception as exc2:  # noqa: BLE001
                    log(f"[×] 回退也失败：{exc2}")

        def do_save_config() -> None:
            ssid = ssid_edit.text().strip() or scan_box.currentText().strip()
            if not ssid:
                QMessageBox.information(dialog, "缺少 SSID", "请先选择或填写 WiFi 名称。")
                return
            payload: dict[str, str] = {"ssid": ssid, "pass": pass_edit.text()}
            if topic_edit.text().strip():
                payload["topic"] = topic_edit.text().strip()
            if sip_edit.text().strip():
                payload["static_ip"] = sip_edit.text().strip()
            if uid_edit.text().strip():
                payload["uid"] = uid_edit.text().strip()
            log(f"[>] POST /config  {urllib.parse.urlencode(payload)}")
            try:
                try:
                    call("/config", payload)
                except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                    # Fallback 到老固件 /api/wifi；注意老固件不支持 topic/static_ip/uid
                    log(f"[·] /config 不可用（{exc}），fallback 旧版 /api/wifi ...")
                    call("/api/wifi", {"ssid": ssid, "pass": pass_edit.text()})
                    log("[√] 旧版接口保存成功（注意：Topic/StaticIP/UID 未写入，新固件支持扩展字段）")
                QMessageBox.information(dialog, "保存成功",
                                        "已写入 WiFi 凭据，模块将在 3 秒后重启并尝试连接。")
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                QMessageBox.warning(dialog, "保存失败", str(exc))

        def do_reconfig() -> None:
            r = QMessageBox.question(dialog, "重置确认",
                                     "会清空 WiFi 配置并让模块重启进入 AP 模式（BanPCBToolXXXX, 12345678）。\n确定？")
            if r != QMessageBox.StandardButton.Yes:
                return
            log("[>] POST /reconfig")
            try:
                call("/reconfig", {})
                QMessageBox.information(dialog, "已重置",
                                        "模块即将重启并进入 AP 模式。\n"
                                        f"SSID：{DEFAULT_DEVICE_NAME}XXXX  密码：{DEFAULT_AP_PASSWORD}")
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                QMessageBox.warning(dialog, "重置失败", str(exc))

        def do_setip() -> None:
            ip = sip_edit.text().strip()
            if not ip:
                QMessageBox.information(dialog, "请填写 IP", "在「静态 IP」里填写地址后再设置。")
                return
            log(f"[>] POST /setip static_ip={ip}")
            try:
                call("/setip", {"static_ip": ip})
                QMessageBox.information(dialog, "已保存静态 IP", "模块将在 1.5s 后重启。")
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                QMessageBox.warning(dialog, "设置失败", str(exc))

        def do_ota() -> None:
            url = ota_edit.text().strip()
            if not url:
                QMessageBox.information(dialog, "缺少 URL", "请先输入 OTA 固件 URL。")
                return
            log(f"[>] POST /api/ota url={url}")
            try:
                call("/api/ota", {"url": url})
            except (OSError, urllib.error.URLError, json.JSONDecodeError) as exc:
                QMessageBox.warning(dialog, "OTA 失败", str(exc))

        def open_browser_update() -> None:
            try:
                import webbrowser
                dev = target_device()
                webbrowser.open(f"http://{dev.ip}:{dev.http_port}/update")
            except (OSError, Exception) as exc:  # noqa: BLE001
                QMessageBox.warning(dialog, "打开失败", str(exc))

        # ===== 信号连接 =====
        scan_box.currentTextChanged.connect(lambda t: ssid_edit.setText(t) if not ssid_edit.text() else None)
        status_btn.clicked.connect(do_status)
        local_scan_btn.clicked.connect(do_local_scan)
        esp_scan_btn.clicked.connect(do_esp_scan)
        save_config_btn.clicked.connect(do_save_config)
        reconfig_btn.clicked.connect(do_reconfig)
        clear_btn.clicked.connect(lambda _=False: (log("[>] POST /api/clear (legacy)"),
                                                    call("/api/clear", {})))
        ap_btn.clicked.connect(lambda _=False: (log("[>] POST /api/ap (开启AP)"),
                                                 call("/api/ap", {})))
        setip_btn.clicked.connect(do_setip)
        restart_btn.clicked.connect(lambda _=False: (log("[>] POST /api/restart"),
                                                      call("/api/restart", {})))
        ota_btn.clicked.connect(do_ota)
        open_update_btn.clicked.connect(open_browser_update)
        close_btn.clicked.connect(dialog.accept)

        # 打开后自动读取一次状态
        QTimer.singleShot(120, do_status)
        dialog.exec()


def run() -> int:
    parser = argparse.ArgumentParser(description="Banux PCB coordinate to G-code desktop tool")
    parser.add_argument("--file", type=Path, help="启动时导入的 XLSX 文件")
    parser.add_argument("--gerber", type=Path, help="启动时导入的 Gerber Paste 文件")
    parser.add_argument("--test", action="store_true", help="导入后输出测试结果并退出")
    parser.add_argument("--screenshot", type=Path, help="保存窗口截图")
    args = parser.parse_args()
    app = QApplication(sys.argv[:1])
    app.setStyle("Fusion")
    window = MainWindow(args.file, args.gerber)
    window.show()
    if args.test or args.screenshot:
        def finish() -> None:
            if args.screenshot:
                args.screenshot.parent.mkdir(parents=True, exist_ok=True)
                window.grab().save(str(args.screenshot))
            if args.test:
                gcode = window.generate_gcode()
                gerber_mode = window.mode_combo.currentIndex() == 1
                print(json.dumps({"file": str(window.current_file) if window.current_file else None,
                                  "gerber": str(window.current_gerber) if window.current_gerber else None,
                                  "mode": "gerber" if gerber_mode else "xlsx",
                                  "sheets": len(window.sheets), "rows": len(window.rows),
                                  "selected": len(window.selected_ids),
                                  "pads": len(window.gerber_pads),
                                  "selected_pads": len(window.selected_pad_ids),
                                  "segments": len(window.pad_dispense_segments()) if gerber_mode else 0,
                                  "gcode_lines": len(gcode.rstrip().splitlines()),
                                  "max_line_bytes": max(len(line.encode("ascii")) for line in gcode.splitlines())},
                                 ensure_ascii=False))
            app.quit()
        QTimer.singleShot(900, finish)
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(run())
